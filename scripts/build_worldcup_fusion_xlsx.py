# -*- coding: utf-8 -*-
"""把 worldcup-fusion-champion.json 渲染成可读 xlsx(归类到带日期子文件夹)。"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = r"D:\football-model-exports\worldcup-fusion-champion.json"
OUTDIR = r"D:\football-model-exports\世界杯多模型融合_2026-06-03"
os.makedirs(OUTDIR, exist_ok=True)
OUT = os.path.join(OUTDIR, "神选-世界杯多模型融合-2026.xlsx")

d = json.load(open(SRC, encoding="utf-8"))
wb = Workbook(); ws = wb.active; ws.title = "多模型融合夺冠概率"

hdr = PatternFill("solid", fgColor="1F4E78")
hf = Font(color="FFFFFF", bold=True, size=11)
edge_pos = PatternFill("solid", fgColor="FCE4D6")
edge_neg = PatternFill("solid", fgColor="DDEBF7")
thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)
center = Alignment(horizontal="center", vertical="center")

cols = ["排名", "球队", "Elo", "本模型(独立)", "自家市场", "Opta超算", "预测市场", "市场共识", "★融合", "分歧(模型-市场)"]
ws.append(cols)
for c in ws[1]:
    c.fill = hdr; c.font = hf; c.alignment = center; c.border = border

for i, r in enumerate(d["rows"], 1):
    row = [i, r["team"], r["elo"], r["eloModel"], r["ourMarket"],
           r.get("opta") or "—", r.get("predMarket") or "—",
           r["mktConsensus"], r["fused"], r["edge"]]
    ws.append(row)
    rr = ws[i + 1]
    for c in rr:
        c.alignment = center; c.border = border
    rr[8].font = Font(bold=True, color="C00000")  # 融合列加粗
    ec = rr[9]
    ec.fill = edge_pos if r["edge"] >= 0 else edge_neg

widths = [6, 12, 7, 12, 10, 10, 10, 10, 9, 16]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(1, i).column_letter].width = w
ws.freeze_panes = "A2"

# 方法论页
ws2 = wb.create_sheet("方法论与诚实说明")
notes = [
    ("方法", d["method"]),
    ("权重", f"wElo={d['weights']['wElo']}  wMkt={d['weights']['wMkt']}"),
    ("源-本模型", d["sources"]["eloModel"]),
    ("源-自家市场", d["sources"]["ourMarket"]),
    ("源-Opta", d["sources"]["opta"]),
    ("源-预测市场", d["sources"]["predMarket"]),
    ("数据日期", d["asOf"]),
    ("", ""),
    ("核心诚实点1", "Opta/预测市场/自家市场三路都含赔率成分,高度相关。直接四路平均=把市场重复计三次。"),
    ("核心诚实点2", "正确做法:三路赔率源塌缩成单一'市场共识',再与本模型唯一独立的 Elo 信号做对数意见池。"),
    ("核心诚实点3", "权重市场偏重(0.65),呼应本项目实证'公开数据打不过市场';融合=市场锚定+模型微调,非'模型最强'。"),
    ("核心诚实点4", "国际赛胜平负命中上限~50-55%(物理天花板),融合不突破它;价值=分布更稳 + 分歧暴露潜在 edge/高估。"),
    ("分歧列读法", "正值=本模型比市场更看好(可能 edge 也可能高估,如西班牙+14/阿根廷+10);负值=本模型看衰(英葡巴德)。"),
    ("", ""),
    ("外部模型冠军投票", "; ".join(f"{v['model']}→{v['champion']}" for v in d["modelPickerVotes"])),
]
for k, v in notes:
    ws2.append([k, v])
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 110
for row in ws2.iter_rows():
    row[0].font = Font(bold=True)
    row[1].alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT)
print("saved:", OUT)
