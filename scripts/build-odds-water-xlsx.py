# -*- coding: utf-8 -*-
"""把 analyze-odds-water-outcomes 的真实回测结论生成速查 xlsx(深紫表头+证据标签)"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = r"C:\Users\Administrator\Desktop\足球推荐\盘口形态实证_2026-06-13"
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, "盘口形态_结果倾向_真实回测速查_2026-06-13.xlsx")

PURPLE = "FF4A148C"; LPURPLE="FFEDE7F6"; GREEN="FFE8F5E9"; RED="FFFFEBEE"; YEL="FFFFF8E1"
hdr_font = Font(name="微软雅黑", bold=True, color="FFFFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor=PURPLE)
title_font = Font(name="微软雅黑", bold=True, color="FFFFFFFF", size=14)
cell_font = Font(name="微软雅黑", size=10)
bold_font = Font(name="微软雅黑", size=10, bold=True)
thin = Side(style="thin", color="FFBDBDBD")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
wrap = Alignment(wrap_text=True, vertical="center")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = Workbook()

def style_header(ws, row, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border

def banner(ws, text, ncol):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
    c = ws.cell(row=1,column=1,value=text); c.font=title_font
    c.fill=PatternFill("solid",fgColor=PURPLE); c.alignment=center
    ws.row_dimensions[1].height=30

# ============ Sheet 1: 速查总表 ============
ws = wb.active; ws.title = "速查总表"
banner(ws, "⚽ 盘口形态 → 结果倾向 · 真实回测速查(5大联赛×7季 12,458场 football-data.co.uk)", 6)
heads = ["盘口形态(你看到的样子)", "实测结果倾向", "实测命中率/过盘率", "样本量", "证据等级", "操作建议"]
ws.append(heads); style_header(ws, 2, 6)

# (形态, 倾向, 数值, n, 证据, 建议, 颜色)
rows = [
 ("【欧赔】收盘超大热(最热方隐含>75%)","按赔率几乎照走","胜率 80.7%","786","✅实测","赔率=最可信胜率估计,大热可信但赔付低",GREEN),
 ("【欧赔】收盘大热(60-75%)","多数兑现","胜率 69.2%","2614","✅实测","校准良好,稳但价值有限",GREEN),
 ("【欧赔】收盘小热(50-60%)","略优于五五","胜率 56.2%","2853","✅实测","接近掷硬币,别重仓",YEL),
 ("【欧赔】最热方<40%(双弱/三方分散)","谁都没把握","最热方仅 39.1%","2125","✅实测","乱战,1X2别强推,考虑双选/弃",RED),
 ("【欧赔】热门被加注(开→收赔率走低)","胜率略升","主56.1%/客57.1%","3091","✅实测","方向真但仅+5pp,且价值已被收盘价吃掉",YEL),
 ("【欧赔】热门退烧(开→收赔率走高)","胜率略降","主51.4%/客47.9%","3141","✅实测","退烧客队尤其要警惕(跌破五成)",RED),
 ("【平局】两队极接近(主客概率差<8pp)","最容易平","打平率 30.6%","1881","✅实测","均势盘平局率远超基准25%,可加保平",GREEN),
 ("【平局】一边倒(强弱差>40pp)","几乎不会平","打平率 18.8%","3868","✅实测","悬殊盘别买平",GREEN),
 ("【亚盘】深盘大热让球(让1球+)且低水(≤1.85)","过盘最差","过盘率 47.8%","439","✅实测","买大热让球上盘长期吃亏,印证‘大热让球<50%’",RED),
 ("【亚盘】深盘大热让球(让1球+)高水(>1.85)","过盘偏低","过盘率 48.4%","3186","✅实测","同上,深盘让球别盲目跟上盘",RED),
 ("【亚盘】浅盘(平半以内)","接近五五","过盘率 50.7%","3374","✅实测","浅盘无明显偏向,看其他因素",YEL),
 ("【盘口变化】升盘(让球加深=机构更看好热门)","热门更会赢但过不了盘","胜率61.6% / 过盘仅45.9%","2005","✅实测","★反直觉:追买升盘热门上盘是-EV,赢球赢不下盘",RED),
 ("【盘口变化】降盘(让球变浅=热门被看淡)","正常","过盘率 50.8%","2089","✅实测","无明显偏向",YEL),
 ("【水位变化】让球线不变·上盘降水(被追)","与玄学相反·几乎无信息","过盘率 51.4%","2316","✅实测","‘降水=必过盘’被证伪,纯水位涨跌≈噪声",YEL),
 ("【水位变化】让球线不变·上盘升水(被抛)","几乎无信息","过盘率 49.8%","2403","✅实测","纯水位升跌对结果无预测力",YEL),
 ("【组合】升盘+升水(传说中的‘诱盘’)","过盘最差的组合","过盘率 45.4%","1505","✅实测","唯一对玄学有弱支持,但主因是‘升盘’非‘水位’",RED),
 ("【大小球】收盘强力大球盘(大球隐含>62%)","多打大球","大球率 69.0%","2297","✅实测","大小球盘校准良好",GREEN),
 ("【大小球】收盘偏小球盘(<45%)","多打小球","大球率 39.6%","2598","✅实测","盘口倾向可信",GREEN),
 ("【大小球】大球被加注(开→收大球赔率走低)","略多大球","大球率 57.9%","3051","✅实测","弱信号,+4.6pp",YEL),
 ("【大小球】大球退烧(转向小球)","略多小球","大球率 48.5%","3696","✅实测","弱信号",YEL),
]
r = 3
for form,trend,val,n,ev,sug,color in rows:
    ws.cell(r,1,form); ws.cell(r,2,trend); ws.cell(r,3,val); ws.cell(r,4,n); ws.cell(r,5,ev); ws.cell(r,6,sug)
    for c in range(1,7):
        cell=ws.cell(r,c); cell.font=cell_font; cell.alignment=wrap; cell.border=border
        if color: cell.fill=PatternFill("solid",fgColor=color)
    ws.cell(r,3).font=bold_font; ws.cell(r,3).alignment=center; ws.cell(r,4).alignment=center; ws.cell(r,5).alignment=center
    r+=1
widths=[34,20,20,8,10,40]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="A3"

# ============ Sheet 2: 校准明细(赔率高低=胜率,市场有效性证据) ============
ws2 = wb.create_sheet("校准明细")
banner(ws2,"📊 收盘赔率分档 → 实际命中率(市场有效性·去水头归一)",4)
ws2.append(["市场类型","收盘隐含概率档","实际命中率","样本量"]); style_header(ws2,2,4)
cal=[
 ("欧赔最热方","A 超大热 >75%","80.7%","786"),
 ("欧赔最热方","B 大热 60-75%","69.2%","2614"),
 ("欧赔最热方","C 小热 50-60%","56.2%","2853"),
 ("欧赔最热方","D 接近五五 40-50%","45.3%","4079"),
 ("欧赔最热方","E 双弱 <40%","39.1%","2125"),
 ("大小球","A 强力大球 >62%","大球69.0%","2297"),
 ("大小球","B 偏大球 55-62%","大球60.5%","2065"),
 ("大小球","C 中性 45-55%","大球50.4%","5496"),
 ("大小球","D 偏小球 <45%","大球39.6%","2598"),
]
r=3
for a,b,c,d in cal:
    ws2.cell(r,1,a);ws2.cell(r,2,b);ws2.cell(r,3,c);ws2.cell(r,4,d)
    for cc in range(1,5):
        cell=ws2.cell(r,cc);cell.font=cell_font;cell.border=border;cell.alignment=center
    r+=1
for i,w in enumerate([14,22,14,10],1): ws2.column_dimensions[get_column_letter(i)].width=w
ws2.freeze_panes="A3"

# ============ Sheet 3: 方法与边界(诚实声明) ============
ws3=wb.create_sheet("方法与边界")
banner(ws3,"🔍 方法、数据源与诚实边界",2)
notes=[
 ("数据源","football-data.co.uk 五大联赛(德甲/英超/法甲/意甲/西甲)×7季(2019/20–2025/26),共12,458场真实赛果"),
 ("赔率口径","开盘=Bet365首盘(B365H/D/A, AHh, B365AHH/AHA, B365>2.5);收盘=赛前最终(B365C*, AHCh, B365CAHH/AHA, B365C>2.5)。1X2按1/赔去水头归一"),
 ("过盘结算","主队让球分数=(主客净球+让球线),四分盘(如-0.25/-0.75)自动拆半结算;走盘计0.5。让球盘设计上本就≈50/50,偏离50%才是信号"),
 ("基准线","全样本:主胜43.1% / 平25.2% / 客胜31.7% / 大球(>2.5)53.3%"),
 ("★核心诚实声明","① 所有数字均为真实历史经验频率,无估计/无兜底;样本<200的桶已标注。② 这些是‘倾向’不是‘必然’,最大偏离也就±5~7pp。③ 公开赔率已编码几乎全部信息,免费数据+简单规则系统性打不过收盘线;盘口形态可减小偏差、辅助避坑,但不保证盈利。④ 唯一可持续edge是速度(赶开盘价收敛前下手)+校准,不是‘看盘猜涨跌’"),
 ("被证伪的玄学","‘降水=看好上盘必过盘’(实测降水过盘51.4%≈升水49.8%,纯水位涨跌是噪声);‘升盘升水=诱盘’只有弱支持且主因是升盘本身。‘逆市场看穿庄家’在更广回测中是亏钱的(分歧越大市场越对)"),
 ("已接入大模型","clv-confidence-gate(背离市场降档)、大热让球过盘惩罚、软赛事平局重校准、开→收漂移信号——本表多数结论已在生产代码中落地"),
]
r=2
for k,v in notes:
    ws3.cell(r,1,k).font=bold_font
    c=ws3.cell(r,2,v); c.font=cell_font; c.alignment=wrap
    ws3.cell(r,1).alignment=Alignment(vertical="top"); ws3.cell(r,1).border=border; c.border=border
    ws3.row_dimensions[r].height=70 if len(v)>60 else 30
    r+=1
ws3.column_dimensions["A"].width=16; ws3.column_dimensions["B"].width=95

wb.save(OUT)
print("SAVED:", OUT)
