# -*- coding: utf-8 -*-
"""check-wc-xlsx —— 竞彩交付xlsx结构/透明度/概率合法性检查器(audit-wc-pipeline S4 调用)。
用法: python scripts/check-wc-xlsx.py <xlsx路径>
输出: 单行JSON {ok, cols, errors[], warnings[], matches:[{home,away}]} 到 stdout。

口径(2026-06-23 用户裁决:主表精简到10列核心版·砍冗余汇总一张主表):
主表10列 = # / 开赛 / 对阵(赛事·情景) / 🎯盘口主推 / 🔥组合触发·三问最可能 /
           让球(让N球后胜平负·主推vs市场) / 比分主推 / 大小球 / 📌研判 / 💰注金。
情报/盘口合理性/返还率/爆冷/决策辅助/半全场/Elo先验等已下沉到专属sheet,主表不再含。
本检查器只验 10列主表;契约冻结在 scripts/delivery-contract.json,改列须重冻并过用户。

仍守的硬口径:让球列必须带 让/受让后胜平负%(模型)vs%(市场)+与胜平负同/不同向标注;
盘口主推必须带方向(p%)或"未开售"原因+信心档徽章;比分主推标来源+≥3档降序;
任何单元格不得出现 undefined/NaN/None 渲染垃圾;深紫表头(FF4A148C)排版标准不被改坏。
"""
import json
import re
import sys

import openpyxl

# 10列核心版关键列(子串匹配真实列头;# 为锚列单独处理)
REQUIRED_HEADERS = [
    "开赛", "对阵", "盘口主推", "组合触发", "让球",
    "比分", "大小球", "研判", "注金",
]
EXPECTED_COLS = 10
PURPLE = "FF4A148C"
GARBAGE = re.compile(r"undefined|\bNaN\b|\bNone\b|\{\{|&lt;span")
TIER = re.compile(r"[🟢🟡🟠⚪]")  # 信心档徽章(已并入盘口主推列)


def main(path):
    errors, warnings, matches = [], [], []
    wb = openpyxl.load_workbook(path)
    if "竞彩完整" not in wb.sheetnames:
        print(json.dumps({"ok": False, "cols": 0, "errors": ["缺'竞彩完整'sheet"], "warnings": [], "matches": []}, ensure_ascii=False))
        return
    if "数据审计" not in wb.sheetnames:
        errors.append("缺'数据审计'sheet(完整性审计是固定标准)")
    ws = wb["竞彩完整"]

    # 自动定位真列头行(标题banner在上方)
    head_row = None
    for r in range(1, min(ws.max_row, 10) + 1):
        vals = [str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1)]
        if vals[0].strip() == "#" and any("对阵" in v for v in vals):
            head_row = r
            break
    if head_row is None:
        errors.append("找不到列头行(首列#+含'对阵')")
        print(json.dumps({"ok": False, "cols": 0, "errors": errors, "warnings": warnings, "matches": []}, ensure_ascii=False))
        return

    headers = [str(ws.cell(head_row, c).value or "") for c in range(1, ws.max_column + 1)]
    # 去前导emoji/空白后按 startswith 精确定位列(组合触发列的长表头含'让球/大小球'子串,
    # 不能用裸 `kw in h` 否则会被它抢走→拿错列)。
    stripped = [re.sub(r"^[^\w一-鿿]+", "", h) for h in headers]
    col = {}
    for kw in REQUIRED_HEADERS:
        idx = next((i + 1 for i, h in enumerate(stripped) if h.startswith(kw)), None)
        if idx is None:  # 退而求其次:含kw但不是组合触发那列
            idx = next((i + 1 for i, h in enumerate(headers) if kw in h and "组合触发" not in h), None)
        if idx is None:
            errors.append(f"缺列: {kw}")
        else:
            col[kw] = idx
    ncols = len([h for h in headers if h])
    if ncols != EXPECTED_COLS:
        errors.append(f"列数{ncols}≠{EXPECTED_COLS}(核心版契约被改;改列须 freeze-delivery-contract --write 重冻+过用户)")

    # 标题banner + 深紫表头(神选标准 FF4A148C)
    if "神选" not in str(ws.cell(1, 1).value or ""):
        warnings.append("标题行无'神选'banner")
    purple = sum(1 for c in range(1, len(headers) + 1)
                 if str(getattr(ws.cell(head_row, c).fill.start_color, "rgb", "")) == PURPLE)
    if purple < ncols * 0.8:
        errors.append(f"表头深紫{PURPLE}仅{purple}/{ncols}格(排版标准被改坏)")

    for r in range(head_row + 1, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        row_name = f"行{r}"
        cells = {kw: str(ws.cell(r, c).value or "") for kw, c in col.items()}
        pair = cells.get("对阵", "")
        # 取首行(第二行起是🏆/情景描述);末尾赛事tag形如"(世界杯·单场)"从尾部剥掉,
        # 避免吃掉队名自带括号(如 刚果(金))
        core = re.sub(r"[\(（][^()（）]*(?:·|联赛|世界杯|单场|杯)[^()（）]*[\)）]\s*$", "", pair.splitlines()[0] if pair else "").strip()
        m = re.match(r"\s*(.+?)\s+vs\s+(.+)$", core)
        if not m:
            errors.append(f"{row_name} 对阵格式异常: {pair[:30]}")
            continue
        home, away = m.group(1).strip(), m.group(2).strip()
        row_name = f"{home}-{away}"
        matches.append({"home": home, "away": away})

        joined = "|".join(cells.values())
        if GARBAGE.search(joined):
            errors.append(f"{row_name} 单元格含渲染垃圾(undefined/NaN/None)")

        # 🎯盘口主推:方向(p%)或"未开售"(带原因) + 信心档徽章(🟢🟡🟠⚪)
        pick = cells.get("盘口主推", "")
        if "未开售" in pick:
            if "(" not in pick and "（" not in pick:
                errors.append(f"{row_name} 盘口主推'未开售'无原因标注")
        else:
            mm = re.search(r"(主胜|平局|客胜)\((\d+)%", pick)
            if not mm:
                errors.append(f"{row_name} 盘口主推缺'方向(p%)': {pick[:30]}")
            else:
                p = int(mm.group(2))
                if not 1 <= p <= 99:
                    errors.append(f"{row_name} 盘口主推概率{p}%越界")
        if not TIER.search(pick):
            errors.append(f"{row_name} 盘口主推缺信心档徽章(🟢🟡🟠⚪): {pick[:30]}")

        # 🔥组合触发·三问最可能:三问齐(胜平负/大小球/让球) 或 诚实标"无欧赔不可算"
        combo = cells.get("组合触发", "")
        if "无欧赔" not in combo and "不可算" not in combo:
            for q in ("胜平负", "大小球", "让球"):
                if f"{q}:" not in combo and f"{q}：" not in combo:
                    errors.append(f"{row_name} 组合触发缺'{q}:'分问: {combo[:40]}")

        # 让球透明度: 模型/市场双数 + 同/不同向标注(独立裁决但必须透明)
        # 合法形态="让2球后胜 53%(模型) vs 60%(市场)…〔让-2〕·与胜平负同向";
        # 预售腿市场缺="…X%(模型)(市场赔率⚠️缺)";绝不许伪造"(市场)"百分数。
        rq = cells.get("让球", "")
        if "缺" in rq and "(模型)" not in rq:
            pass  # 让球真实裁决整列缺(诚实形态)
        else:
            market_ok = "(市场)" in rq or re.search(r"市场[^〕)]*⚠?️?缺", rq)
            if "(模型)" not in rq or not market_ok:
                errors.append(f"{row_name} 让球列缺'让/受让后胜平负%(模型)vs%(市场)'透明双数: {rq[:30]}")
            if "与胜平负" not in rq:
                errors.append(f"{row_name} 让球列缺与胜平负同/不同向标注")

        # 比分主推: 来源标注(盘口主推✅ 或 🔶模型) + ≥3个 比分(p%);只查前3档降序
        score = cells.get("比分", "")
        sc = [int(x) for x in re.findall(r"\d+-\d+\((\d+)%\)", score)]
        if "主推" not in score and "🔶" not in score:
            errors.append(f"{row_name} 比分列缺'主推'来源标注")
        if len(sc) < 3:
            errors.append(f"{row_name} 比分主推不足3档: {score[:30]}")
        elif any(sc[i] < sc[i + 1] for i in range(2)):
            errors.append(f"{row_name} 比分主推前3档非降序: {sc[:3]}")

        # 大小球: 大NN%/小NN% 真盘 或 诚实标缺
        ou = cells.get("大小球", "")
        oum = re.search(r"大(\d+)%/小(\d+)%", ou)
        if oum:
            s = int(oum.group(1)) + int(oum.group(2))
            if not 98 <= s <= 102:
                errors.append(f"{row_name} 大小球 大+小={s}≠100")
        elif "缺" not in ou and "⚠️" not in ou and "未" not in ou:
            errors.append(f"{row_name} 大小球列无'大%/小%'也无缺数标注: {ou[:30]}")

        # 审计轨迹列非空(研判/注金)
        for kw in ("研判", "注金"):
            if not cells.get(kw, "").strip():
                errors.append(f"{row_name} '{kw}'列为空(审计轨迹断)")

    if not matches:
        errors.append("竞彩完整sheet无数据行")
    print(json.dumps({"ok": not errors, "cols": len([h for h in headers if h]),
                      "errors": errors, "warnings": warnings, "matches": matches}, ensure_ascii=False))


if __name__ == "__main__":
    main(sys.argv[1])
