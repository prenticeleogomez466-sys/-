"""
2026-05-28 足球大模型综合评分报告(顶级团队视角).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT = r"C:\Users\Administrator\Desktop\2026-05-28 足球大模型评分报告.xlsx"

# ───── 评分细则 ─────
dimensions = [
    {
        "name": "数据层",
        "max": 20,
        "score": 13,
        "items": [
            ("数据源数量", 5, 4, "10+ 源(sporttery/500/新浪/fotmob/ClubElo/football-data/GDELT/Open-Meteo/OpenFootball/Understat)", "缺中超、巴甲、阿甲专属源"),
            ("数据广度", 5, 3, "覆盖五大联赛 + 解放者杯 + 国际赛 + 中超 14 场", "缺 event-level 位置数据、StatsBomb/Opta 付费源"),
            ("数据深度", 5, 3, "fotmob 总 xG + Understat 每 shot xG + 比分赔率 + 半全场赔率", "缺球员市值(Transfermarkt)、缺压力/触球热区"),
            ("数据稳定性", 5, 3, "WAF 多重重试 + UA 池 + 5 级救援 + Playwright Chrome 备份", "依赖公开页面,庄家反爬升级时需补)"),
        ]
    },
    {
        "name": "模型层",
        "max": 25,
        "score": 18,
        "items": [
            ("基础统计模型", 8, 7, "Dixon-Coles 泊松(含扩展 tau)+ Bivariate Poisson + Hierarchical Poisson", "缺 Bayesian state-space + MCMC 完整版"),
            ("球队评级", 6, 6, "Elo + Pi-ratings + Massey + Colley = 4 套评级", "—(满分,行业最佳)"),
            ("集成学习", 6, 3, "线性逻辑回归 stacker(LR softmax)+ Ratings ensemble", "**缺真 XGBoost / LightGBM / CatBoost / Neural Network**"),
            ("校准与冷启动", 5, 2, "Isotonic 校准 + cold-start 联赛先验 + signal-weights profile", "新评级模块**未接入 daily 流水线**,样本不够时优势浪费"),
        ]
    },
    {
        "name": "输出层",
        "max": 15,
        "score": 13,
        "items": [
            ("玩法覆盖", 6, 6, "胜负+让球+比分+半全场+大小球(5档)+单双+上半场+双胜彩+比分组+总进球+串关", "—(满分)"),
            ("决策标签", 5, 4, "EV 标签 + verdict(strong-value/value/fair/negative-ev)+ 半凯利仓位 + 联合 EV", "缺 dutching / hedge / arbitrage 提示"),
            ("串关组合", 4, 3, "二/三/四/五串一全 EV 排序 + 含避坑场过滤 + 仓位建议", "缺自动化 combo 优化器(凯利分配多腿仓位)"),
        ]
    },
    {
        "name": "闭环系统",
        "max": 15,
        "score": 11,
        "items": [
            ("自动化抓取", 5, 5, "schtasks 03:00 daily + 11:00 recap + 09:32 health,Playwright 兜底", "—(满分)"),
            ("复盘校准", 5, 4, "daily-recap + evolution-backtest + RPS 评估 + calibration-profile + signal-weights", "Pi/Massey/Colley/Hier 未参与 backtest 计算 RPS"),
            ("自我演化", 5, 2, "isotonic + favorite-longshot 冷启动 prior + signal weights 自适应", "Linear stacker / 新评级**未真正集成进 prediction pipeline**"),
        ]
    },
    {
        "name": "工程稳定性",
        "max": 10,
        "score": 9,
        "items": [
            ("测试覆盖", 4, 4, "118 个单测,11 个 test 文件,覆盖核心算法 + 数据解析 + 边界条件", "—(满分)"),
            ("容错降级", 3, 3, "HTTP 567 重试 + UA 池 + 多源救援 + cold-start fallback + automation host fix", "—(满分)"),
            ("可维护性", 3, 2, "模块化 src/ 单一职责 + 中文注释 + memory 持久知识", "缺 CI/CD + 集成测试"),
        ]
    },
    {
        "name": "决策辅助",
        "max": 10,
        "score": 8,
        "items": [
            ("透明度", 4, 4, "市场结构解读 + 让球盘解读 + 半全场赔率分析 + 比分赔率结构", "—(满分)"),
            ("解释性", 3, 2, "每场附 pick_reason / score_reason / hf_reason + 爆冷场景", "目前手动撰写,未自动生成"),
            ("风险提示", 3, 2, "爆冷指数(低/中/中高/高)+ 阵容备注 + EV verdict + vig 警告", "无凯利破产警告 + 无连败回撤提醒"),
        ]
    },
    {
        "name": "用户体验",
        "max": 5,
        "score": 5,
        "items": [
            ("输出格式", 2, 2, "xlsx + 单 sheet + 微软雅黑 + 冻结表头 + 颜色色阶 + 自动列宽", "—(满分)"),
            ("可读性", 2, 2, "建议色阶(绿/黄/橙/红)+ 直方图 + 概率百分比 + 中文方向描述", "—(满分)"),
            ("及时性", 1, 1, "Playwright 实时抓取 + 桌面直接落盘 + memory 持久化偏好", "—(满分)"),
        ]
    }
]

# ───── 路径到 90+ ─────
upgrade_paths = [
    ("把 D 档算法真正接入 daily 流水线", "+5", "Pi/Massey/Colley/Hier/Bivariate 都集成到 prediction-engine,buildEnsemblePrediction 投票", "1-2h", "立刻"),
    ("Python + ONNX 真 XGBoost stacking", "+4", "用 ledger 历史 + OpenFootball + Understat 训练 XGBoost / LightGBM / CatBoost,ONNX 落盘,Node 推理", "1-2 天", "短期"),
    ("Bayesian state-space DC(动态球队强度)", "+2", "重写 DC 引擎,加入时间动态(球队强度随赛季演化)", "1-2 天", "中期"),
    ("中超 + 巴甲 + 阿甲专属数据源", "+2", "找 GitHub 中超数据仓库,接入 advanced-data-runner", "半天", "短期"),
    ("Transfermarkt 球员市值特征", "+2", "dcaribou/transfermarkt-datasets CSV,作为强度先验", "1 天", "中期"),
    ("付费 StatsBomb / Opta event-level", "+2", "**需用户付费**(StatsBomb open data 有限免)", "看授权", "远期"),
    ("Neural Network(tfjs MLP)", "+1", "5 隐层 + Dropout,跟 XGBoost stacker 一起 ensemble", "1 天", "中期"),
    ("Ledger 样本积累到 300+", "+1", "时间问题,backtest 自动样本越多越准", "等几个月", "自然演化"),
]

# ───── 样式 ─────
TITLE_FONT = Font(name="微软雅黑", bold=True, size=18, color="FFFFFF")
TITLE_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="305496")
SECTION_FONT = Font(name="微软雅黑", bold=True, size=14, color="FFFFFF")
SECTION_FILL = PatternFill("solid", start_color="2E75B6")
DEFAULT_FONT = Font(name="微软雅黑", size=11)
BOLD_FONT = Font(name="微软雅黑", bold=True, size=11)
SMALL_FONT = Font(name="微软雅黑", size=10)
ALT_FILL = PatternFill("solid", start_color="F2F7FC")
THIN = Side(border_style="thin", color="9CC3E5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

def grade_color(score, max_score):
    pct = score / max_score
    if pct >= 0.9: return "00B050"  # 优 绿
    if pct >= 0.75: return "92D050" # 良 浅绿
    if pct >= 0.6: return "FFC000"  # 中 橙
    return "C00000"  # 差 红

def grade_label(score, max_score):
    pct = score / max_score
    if pct >= 0.9: return "优"
    if pct >= 0.75: return "良"
    if pct >= 0.6: return "中"
    return "差"

wb = Workbook()
s = wb.active
s.title = "足球大模型评分"

# 标题
total = sum(d["score"] for d in dimensions)
target = 90
gap = target - total

s.cell(row=1, column=1, value="🏆 足球大模型综合评分报告(2026-05-28)")
s.cell(row=1, column=1).font = TITLE_FONT
s.cell(row=1, column=1).fill = TITLE_FILL
s.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
s.merge_cells("A1:F1")
s.row_dimensions[1].height = 38

s.cell(row=2, column=1, value=f"评估视角: 世界顶级足球分析团队  |  总分: {total} / 100  |  目标: 90+  |  差距: {gap} 分  |  评级: {'A 级' if total >= 90 else 'B+ 级' if total >= 80 else 'B 级' if total >= 70 else 'C 级'}")
s.cell(row=2, column=1).font = Font(name="微软雅黑", italic=True, size=11, color="595959")
s.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
s.merge_cells("A2:F2")
s.row_dimensions[2].height = 22

# 区段 1:总分概览
s.cell(row=4, column=1, value="一、七维度总分概览")
s.cell(row=4, column=1).font = SECTION_FONT
s.cell(row=4, column=1).fill = SECTION_FILL
s.cell(row=4, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
s.merge_cells("A4:F4")
s.row_dimensions[4].height = 26

headers = ["维度", "满分", "得分", "占比", "评级", "主要缺口"]
for j, h in enumerate(headers, 1):
    c = s.cell(row=5, column=j, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
s.row_dimensions[5].height = 26

main_gaps = {
    "数据层": "缺 event-level + Transfermarkt + 中超",
    "模型层": "无真 XGBoost/NN,新评级未集成",
    "输出层": "缺 dutching / 凯利组合优化",
    "闭环系统": "新评级未参与 daily/backtest",
    "工程稳定性": "缺 CI/CD",
    "决策辅助": "解释/风险提示偏手动",
    "用户体验": "—"
}

for i, d in enumerate(dimensions, 6):
    alt = (i % 2 == 0)
    pct = d["score"] / d["max"]
    label = grade_label(d["score"], d["max"])
    color = grade_color(d["score"], d["max"])
    s.cell(row=i, column=1, value=d["name"]).font = BOLD_FONT
    s.cell(row=i, column=2, value=d["max"])
    s.cell(row=i, column=3, value=d["score"])
    s.cell(row=i, column=4, value=pct).number_format = "0.0%"
    s.cell(row=i, column=5, value=label).fill = PatternFill("solid", start_color=color)
    s.cell(row=i, column=5).font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    s.cell(row=i, column=6, value=main_gaps.get(d["name"], ""))
    for j in range(1, 7):
        c = s.cell(row=i, column=j)
        c.alignment = CENTER if j != 6 else LEFT_WRAP
        c.border = BORDER
        if alt and c.fill.start_color.rgb is None: c.fill = ALT_FILL
        if j == 1: c.font = BOLD_FONT
        elif j != 5: c.font = DEFAULT_FONT
    s.row_dimensions[i].height = 26

# 总分行
i = 6 + len(dimensions)
s.cell(row=i, column=1, value="合计").font = Font(name="微软雅黑", bold=True, size=12)
s.cell(row=i, column=2, value=100).font = BOLD_FONT
s.cell(row=i, column=3, value=total).font = Font(name="微软雅黑", bold=True, size=14, color="C00000")
s.cell(row=i, column=4, value=total/100).number_format = "0.0%"
s.cell(row=i, column=4).font = BOLD_FONT
s.cell(row=i, column=5, value=f"距 90 差 {gap}")
s.cell(row=i, column=5).fill = PatternFill("solid", start_color="ED7D31")
s.cell(row=i, column=5).font = Font(name="微软雅黑", bold=True, color="FFFFFF")
s.cell(row=i, column=6, value="见下方升级路径")
for j in range(1, 7):
    s.cell(row=i, column=j).alignment = CENTER if j != 6 else LEFT_WRAP
    s.cell(row=i, column=j).border = BORDER
s.row_dimensions[i].height = 28

# 区段 2:逐项细则
row = i + 2
s.cell(row=row, column=1, value="二、逐项细则(每个维度内的具体得分)")
s.cell(row=row, column=1).font = SECTION_FONT
s.cell(row=row, column=1).fill = SECTION_FILL
s.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
s.merge_cells(f"A{row}:F{row}")
s.row_dimensions[row].height = 26
row += 1

sub_headers = ["维度", "细项", "满分", "得分", "已做到", "未做到/缺口"]
for j, h in enumerate(sub_headers, 1):
    c = s.cell(row=row, column=j, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
s.row_dimensions[row].height = 24
row += 1

for d in dimensions:
    first_in_dim = True
    for item in d["items"]:
        name, max_pts, got, done, missing = item
        alt = (row % 2 == 0)
        if first_in_dim:
            c = s.cell(row=row, column=1, value=d["name"])
            c.font = BOLD_FONT
            first_in_dim = False
        else:
            c = s.cell(row=row, column=1, value="")
        s.cell(row=row, column=2, value=name).font = DEFAULT_FONT
        s.cell(row=row, column=3, value=max_pts)
        gc = grade_color(got, max_pts)
        s.cell(row=row, column=4, value=got).fill = PatternFill("solid", start_color=gc)
        s.cell(row=row, column=4).font = Font(name="微软雅黑", bold=True, color="FFFFFF")
        s.cell(row=row, column=5, value=done)
        s.cell(row=row, column=6, value=missing)
        for j in range(1, 7):
            cell = s.cell(row=row, column=j)
            cell.alignment = CENTER if j in [1, 2, 3, 4] else LEFT_WRAP
            cell.border = BORDER
            if alt and cell.fill.start_color.rgb is None: cell.fill = ALT_FILL
        s.row_dimensions[row].height = 38
        row += 1

# 区段 3:90+ 路径
row += 1
s.cell(row=row, column=1, value="三、达到 90+ 的升级路径")
s.cell(row=row, column=1).font = SECTION_FONT
s.cell(row=row, column=1).fill = SECTION_FILL
s.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
s.merge_cells(f"A{row}:F{row}")
s.row_dimensions[row].height = 26
row += 1

up_headers = ["#", "升级项", "可加分", "具体动作", "工作量", "优先级"]
for j, h in enumerate(up_headers, 1):
    c = s.cell(row=row, column=j, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
s.row_dimensions[row].height = 24
row += 1

priority_colors = {"立刻": "00B050", "短期": "92D050", "中期": "FFC000", "远期": "ED7D31", "自然演化": "BFBFBF"}
running_total = total
for k, (name, gain, action, effort, priority) in enumerate(upgrade_paths, 1):
    alt = (row % 2 == 0)
    s.cell(row=row, column=1, value=k)
    s.cell(row=row, column=2, value=name)
    s.cell(row=row, column=3, value=gain).font = Font(name="微软雅黑", bold=True, color="00B050")
    s.cell(row=row, column=4, value=action)
    s.cell(row=row, column=5, value=effort)
    pc = priority_colors.get(priority, "808080")
    s.cell(row=row, column=6, value=priority).fill = PatternFill("solid", start_color=pc)
    s.cell(row=row, column=6).font = Font(name="微软雅黑", bold=True, color="FFFFFF")
    for j in range(1, 7):
        c = s.cell(row=row, column=j)
        c.alignment = CENTER if j in [1, 3, 5, 6] else LEFT_WRAP
        c.border = BORDER
        if alt and c.fill.start_color.rgb is None: c.fill = ALT_FILL
        if j == 2: c.font = BOLD_FONT
        elif j not in [3, 6]: c.font = DEFAULT_FONT
    s.row_dimensions[row].height = 30
    row += 1

# 优先路径汇总
row += 1
s.cell(row=row, column=1, value="🎯 推荐路径")
s.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=12, color="C00000")
s.cell(row=row, column=2, value="按 #1 (D 档接入 daily) + #2 (XGBoost) + #4 (中超数据) = +11 分 → 88 分(接近 90)")
s.cell(row=row, column=2).font = DEFAULT_FONT
s.cell(row=row, column=2).alignment = LEFT_WRAP
s.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
s.row_dimensions[row].height = 22
row += 1
s.cell(row=row, column=2, value="再做 #3 (Bayesian state-space) + #5 (Transfermarkt) = +4 分 → 92 分(A 级)")
s.cell(row=row, column=2).font = DEFAULT_FONT
s.cell(row=row, column=2).alignment = LEFT_WRAP
s.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
s.row_dimensions[row].height = 22
row += 1

# 评分原则说明
row += 2
s.cell(row=row, column=1, value="四、评分原则(顶级团队视角)")
s.cell(row=row, column=1).font = SECTION_FONT
s.cell(row=row, column=1).fill = SECTION_FILL
s.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
s.merge_cells(f"A{row}:F{row}")
s.row_dimensions[row].height = 26
row += 1

principles = [
    ("命中率不直接打分", "顶级商业模型胜负彩上限 56-58%(数学物理),所以「命中率分数」是误导。我们打综合能力分:模型工程质量 + 决策辅助 + 闭环 + UX"),
    ("数据广度和深度并重", "数据源数量(广度)和单源精度(深度)都重要;Understat 每 shot xG > fotmob 总 xG"),
    ("模型多样性 > 单模型精度", "Ensemble 6 个模型比单 DC 强,因为不同模型捕捉不同特征"),
    ("闭环 > 静态最优", "能自我演化的 77 分模型 > 静态的 85 分模型;ledger + backtest + calibration profile 是核心资产"),
    ("透明 > 黑盒", "顶级团队为客户工作,必须解释「为什么」,所以爆冷指数 + 市场解读 + 阵容备注是必需"),
    ("付费数据不算扣分", "如果用户不愿付费 StatsBomb/Opta,我们不能因此扣他分;但他们能加多少分,如实告诉他")
]
for label, txt in principles:
    s.cell(row=row, column=1, value=label).font = Font(name="微软雅黑", bold=True, size=10)
    s.cell(row=row, column=1).fill = ALT_FILL
    s.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    s.cell(row=row, column=2, value=txt).font = SMALL_FONT
    s.cell(row=row, column=2).alignment = LEFT_WRAP
    s.cell(row=row, column=2).fill = ALT_FILL
    s.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    for j in range(1, 7):
        s.cell(row=row, column=j).border = BORDER
    s.row_dimensions[row].height = 32
    row += 1

# 列宽
widths = {"A": 14, "B": 22, "C": 8, "D": 38, "E": 16, "F": 28}
for letter, w in widths.items():
    s.column_dimensions[letter].width = w

s.freeze_panes = "A6"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Total: {total}/100, target 90+, gap {gap} points")
