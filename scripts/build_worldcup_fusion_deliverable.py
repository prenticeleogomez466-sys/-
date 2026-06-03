# -*- coding: utf-8 -*-
"""世界杯融合交付物:xlsx(足球数据分析库)+ 手机网页。读 exports 的官方表+γ JSON。"""
import json, os, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXP = r"D:\football-model-exports"
fus = json.load(open(os.path.join(EXP, "worldcup-fusion-champion.json"), encoding="utf-8"))
sup = json.load(open(os.path.join(EXP, "worldcup-supercomputer.json"), encoding="utf-8"))
ASOF = fus.get("asOf", str(datetime.date.today()))

xdir = r"D:\足球数据分析库\世界杯2026"
os.makedirs(xdir, exist_ok=True)
xpath = os.path.join(xdir, f"神选-世界杯融合-{ASOF}.xlsx")

hf = Font(bold=True, color="FFFFFF"); hp = PatternFill("solid", fgColor="1F4E78")
gold = PatternFill("solid", fgColor="FFF2CC")
cen = Alignment(horizontal="center", vertical="center"); left = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin", color="D9D9D9"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
title_f = Font(bold=True, size=13, color="1F4E78"); note_f = Font(italic=True, size=9, color="808080")

wb = Workbook()

# ── Sheet1 融合冠军榜 ──
ws = wb.active; ws.title = "融合冠军榜"
ws["A1"] = f"2026 世界杯·多模型融合夺冠概率(官方对阵表+Rue-Salvesen γ,as of {ASOF})"
ws["A1"].font = title_f; ws.merge_cells("A1:J1")
ws["A2"] = "方法:对数意见池 = 本模型 Elo独立信号 ⊕ 市场共识[Opta超算+预测市场+自家去抽水],权重 wMkt0.65/wElo0.35;edge=模型−市场共识(正=模型更看好,按实证多为高估)。"
ws["A2"].font = note_f; ws.merge_cells("A2:J2")
hdr = ["排名", "球队", "Elo", "模型%", "自家市场%", "Opta%", "预测市场%", "市场共识%", "融合%", "分歧edge"]
ws.append([]); ws.append(hdr)
for c in ws[4]: c.font = hf; c.fill = hp; c.alignment = cen; c.border = bd
rows = sorted(fus["rows"], key=lambda r: -r["fused"])
for i, r in enumerate(rows, 1):
    ws.append([i, r["team"], r["elo"], r["eloModel"], r["ourMarket"], r.get("opta", ""), r.get("predMarket", ""), r["mktConsensus"], r["fused"], r["edge"]])
    rr = ws.max_row
    for c in ws[rr]:
        c.border = bd; c.alignment = cen if c.column != 2 else left
    if i <= 3:
        for c in ws[rr]: c.fill = gold
for col, w in zip("ABCDEFGHIJ", [5, 12, 6, 8, 10, 8, 10, 10, 8, 9]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ── Sheet2 超算各阶段晋级率 ──
ws2 = wb.create_sheet("超算各阶段晋级率")
ws2["A1"] = f"2026 世界杯·超算各阶段到达概率(N={sup['n']} 蒙特卡洛,FIFA官方对阵表+γ,审计{'✓' if sup['audit']['ok'] else '✗'})"
ws2["A1"].font = title_f; ws2.merge_cells("A1:K1")
hdr2 = ["排名", "球队", "Elo", "出线%", "16强%", "8强%", "4强%", "决赛%", "夺冠模型%", "市场%", "混合%"]
ws2.append([]); ws2.append(hdr2)
for c in ws2[3]: c.font = hf; c.fill = hp; c.alignment = cen; c.border = bd
srows = sorted(sup["rows"], key=lambda r: -r["blend"])
P = lambda x: round(x * 100, 1)
for i, r in enumerate(srows, 1):
    ws2.append([i, r["team"], r["elo"], P(r["advance"]), P(r["r16"]), P(r["qf"]), P(r["sf"]), P(r["final"]), P(r["champion"]), P(r["market"]), P(r["blend"])])
    rr = ws2.max_row
    for c in ws2[rr]:
        c.border = bd; c.alignment = cen if c.column != 2 else left
    if i <= 3:
        for c in ws2[rr]: c.fill = gold
for col, w in zip("ABCDEFGHIJK", [5, 12, 6, 8, 8, 8, 8, 8, 11, 8, 8]):
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = "A4"
wb.save(xpath)
print("XLSX:", xpath)

# ── 手机网页 ──
def tr(i, r, gold3):
    cls = ' class="g"' if i <= gold3 else ""
    return f"<tr{cls}><td>{i}</td><td class='t'>{r['team']}</td><td>{r['elo']}</td><td>{r['eloModel']:.1f}</td><td>{r['mktConsensus']:.1f}</td><td><b>{r['fused']:.1f}</b></td><td>{r['edge']:+.1f}</td></tr>"
frows = "".join(tr(i, r, 3) for i, r in enumerate(rows, 1))
def tr2(i, r):
    cls = ' class="g"' if i <= 3 else ""
    return f"<tr{cls}><td>{i}</td><td class='t'>{r['team']}</td><td>{P(r['advance'])}</td><td>{P(r['qf'])}</td><td>{P(r['sf'])}</td><td>{P(r['final'])}</td><td><b>{P(r['blend'])}</b></td></tr>"
srows_html = "".join(tr2(i, r) for i, r in enumerate(srows, 1))
html = f"""<!doctype html><html lang=zh><head><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1">
<title>2026世界杯·神选融合</title>
<style>
body{{font-family:-apple-system,system-ui,sans-serif;margin:0;background:#0f1115;color:#e8e8e8}}
.wrap{{max-width:780px;margin:0 auto;padding:14px}}
h1{{font-size:19px;margin:8px 0 2px}} .sub{{color:#9aa0a6;font-size:12px;margin-bottom:12px;line-height:1.5}}
h2{{font-size:15px;margin:20px 0 6px;color:#7fb3ff}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th,td{{padding:6px 4px;text-align:center;border-bottom:1px solid #23262e}}
th{{background:#1b2330;color:#cfe0ff;position:sticky;top:0}}
td.t{{text-align:left;font-weight:600}} tr.g td{{background:#2a2410}}
b{{color:#ffd479}} .foot{{color:#6b7178;font-size:11px;margin:16px 0;line-height:1.6}}
</style></head><body><div class=wrap>
<h1>⚽ 2026 世界杯 · 神选融合预测</h1>
<div class=sub>官方对阵表(R32真实位次+第三名495分配)+ Rue-Salvesen γ收缩 + 多模型对数意见池融合(本模型⊕Opta⊕预测市场,市场锚0.65)。as of {ASOF}。</div>
<h2>🏆 夺冠概率(融合)</h2>
<table><tr><th>#</th><th>球队</th><th>Elo</th><th>模型%</th><th>市场共识%</th><th>融合%</th><th>分歧</th></tr>{frows}</table>
<h2>📊 各阶段晋级率(超算)</h2>
<table><tr><th>#</th><th>球队</th><th>出线%</th><th>8强%</th><th>4强%</th><th>决赛%</th><th>夺冠混合%</th></tr>{srows_html}</table>
<div class=foot>诚实边界:国际赛胜平负命中天花板 ~50-55%(爆冷常态),融合不破天花板,价值在分布更稳+分歧暴露(分歧大处实证多为模型高估,故市场锚定)。点球50/50(学界)。数据源:eloratings.net / Opta / 预测市场,均免费。审计:夺冠和100%·出线和32·单调✓。</div>
</div></body></html>"""
hpath = r"D:\Temp\webshare_lingdao\worldcup.html"
open(hpath, "w", encoding="utf-8").write(html)
print("HTML:", hpath)
print("TOP3 融合:", " / ".join(f"{r['team']}{r['fused']:.1f}%" for r in rows[:3]))
