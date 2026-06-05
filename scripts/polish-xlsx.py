"""
神选气质 xlsx 美化:
- 配色:深紫(#4A148C 神圣感) + 金色(#F9A825 高级感) + 玄黑(#1A1A1A) + 浅紫底(#F3E5F5)
- 顶部加标题区:"⚡ 神选 · 竞彩推荐 · YYYY-MM-DD" 合并单元格 + 金色字
- 表头:深紫底 + 金色字加粗 14 号 + 居中 + 行高 38
- 数据行:奇偶交替白/浅紫底 + 12 号字 + 行高 28
- 胜平负列条件染色:主胜 浅绿 / 平局 金黄 / 客胜 浅红(配合神选金色调)
- 胆码列:深金色高亮加粗(神级推荐)
- 信心列:🟢 加粗深紫,🔴 加粗深红
- 底部加签名行"⚡ Claude 神选 · 独立大模型 · 自主推断"
"""
import sys, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if len(sys.argv) < 2:
    print("用法: python polish-xlsx.py <path>", file=sys.stderr)
    sys.exit(1)

PATH = sys.argv[1]
DATE = re.search(r'(\d{4}-\d{2}-\d{2})', PATH)
DATE_STR = DATE.group(1) if DATE else ""

wb = load_workbook(PATH)

# === 神选配色 ===
DEEP_PURPLE   = "FF4A148C"     # 深紫 — 神圣
GOLD          = "FFF9A825"     # 金色 — 高级
LIGHT_PURPLE  = "FFF3E5F5"     # 浅紫 — 偶数行底
PURE_WHITE    = "FFFFFFFF"
SHEN_BLACK    = "FF1A1A1A"     # 玄黑 — 主文字
SOFT_GREEN    = "FFC8E6C9"     # 主胜 浅绿
SOFT_YELLOW   = "FFFFF59D"     # 平局 金黄
SOFT_RED      = "FFFFCDD2"     # 客胜 浅红
BANKER_GOLD   = "FFFFB300"     # 胆码 深金
DOUBLE_BLUE   = "FFE1F5FE"     # 双选 浅蓝
TRIPLE_GRAY   = "FFF5F5F5"     # 全选 浅灰

# === 字体 ===
TITLE_FONT  = Font(name="Microsoft YaHei", size=18, bold=True, color=GOLD)
HEADER_FONT = Font(name="Microsoft YaHei", size=13, bold=True, color=GOLD)
DATA_FONT   = Font(name="Microsoft YaHei", size=11, color=SHEN_BLACK)
DATA_FONT_B = Font(name="Microsoft YaHei", size=11, bold=True, color=DEEP_PURPLE)
HIGH_CONF   = Font(name="Microsoft YaHei", size=11, bold=True, color=DEEP_PURPLE)
LOW_CONF    = Font(name="Microsoft YaHei", size=11, color="FFC62828")
BANKER_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="FF8B4513")
SIGN_FONT   = Font(name="Microsoft YaHei", size=10, italic=True, color=DEEP_PURPLE)

# === 填充 ===
TITLE_FILL  = PatternFill("solid", fgColor=DEEP_PURPLE)
HEADER_FILL = PatternFill("solid", fgColor=DEEP_PURPLE)
EVEN_FILL   = PatternFill("solid", fgColor=LIGHT_PURPLE)
ODD_FILL    = PatternFill("solid", fgColor=PURE_WHITE)
HOME_FILL   = PatternFill("solid", fgColor=SOFT_GREEN)
DRAW_FILL   = PatternFill("solid", fgColor=SOFT_YELLOW)
AWAY_FILL   = PatternFill("solid", fgColor=SOFT_RED)
BANKER_FILL = PatternFill("solid", fgColor=BANKER_GOLD)
DOUBLE_FILL = PatternFill("solid", fgColor=DOUBLE_BLUE)
TRIPLE_FILL = PatternFill("solid", fgColor=TRIPLE_GRAY)
SIGN_FILL   = PatternFill("solid", fgColor=LIGHT_PURPLE)

# === 对齐 + 边框 ===
CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
GOLD_SIDE = Side(style="thin", color=GOLD)
PURPLE_SIDE = Side(style="thin", color="FFCE93D8")
HEADER_BORDER = Border(top=GOLD_SIDE, bottom=GOLD_SIDE, left=GOLD_SIDE, right=GOLD_SIDE)
DATA_BORDER = Border(top=PURPLE_SIDE, bottom=PURPLE_SIDE, left=PURPLE_SIDE, right=PURPLE_SIDE)


def col_width_for(header, idx):
    h = str(header or "")
    if any(k in h for k in ["选择理由", "理由", "说明", "融合判断要点", "narrative"]): return 62
    if any(k in h for k in ["对阵", "比赛"]): return 30
    if "概率分布" in h or "概率(主/平/客)" in h or "主/平/客" in h: return 30
    if "信心 · 分级" in h or "信心 ·" in h: return 36
    if "让胜负平" in h or "让球" in h: return 24
    if any(k in h for k in ["半全场", "比分"]): return 20
    if any(k in h for k in ["赛事类型", "爆冷"]): return 20
    if "胜负平" in h or "胜平负" in h: return 16
    if "开赛" in h: return 14
    if "覆盖" in h or "单式" in h: return 18
    if h == "赛事": return 14
    if h in ("序", "场次", "类型"): return 10
    if idx <= 1: return 12
    if idx <= 4: return 20
    return 22


for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws.max_row < 1: continue
    max_col = ws.max_column
    last_col_letter = get_column_letter(max_col)

    headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]

    # === 插入顶部标题行 ===
    ws.insert_rows(1)
    title_text = f"⚡ 神选 · {sheet_name} · {DATE_STR}"
    ws.cell(1, 1, value=title_text)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(1, 1)
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 44

    # === 表头 ===
    ws.row_dimensions[2].height = 38
    for c in range(1, max_col + 1):
        cell = ws.cell(2, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(c)].width = col_width_for(headers[c-1], c-1)

    # 胜负平方向染色:竞彩极简表里「胜负平」与「让胜负平」两列都按各自方向上色,
    #   同向 → 同色,用户一眼即可看出四列方向一致(不一致会立刻露馅)。
    wld_cols = [i+1 for i, h in enumerate(headers) if "胜负平" in str(h) or "胜平负" in str(h) or str(h) == "单式"]
    wld_col = wld_cols[0] if wld_cols else None
    conf_col = next((i+1 for i, h in enumerate(headers) if "信心" in str(h)), None)
    type_col = next((i+1 for i, h in enumerate(headers) if h == "类型"), None)
    long_cols = [i+1 for i, h in enumerate(headers) if any(k in str(h or "") for k in ["选择理由", "比赛", "对阵", "说明", "融合判断要点", "evidence", "Evidence", "narrative"])]

    # === 数据行 ===
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 28
        is_even = ((r - 3) % 2 == 0)
        row_fill = EVEN_FILL if is_even else ODD_FILL
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.font = DATA_FONT
            cell.border = DATA_BORDER
            cell.fill = row_fill
            cell.alignment = LEFT if c in long_cols else CENTER

        # wld 方向染色(胜负平 + 让胜负平 两列各按自身方向)
        for wc in wld_cols:
            v = str(ws.cell(r, wc).value or "")
            if "主胜" in v: ws.cell(r, wc).fill = HOME_FILL
            elif "客胜" in v: ws.cell(r, wc).fill = AWAY_FILL
            elif "平局" in v or "走盘" in v or v.strip() == "平": ws.cell(r, wc).fill = DRAW_FILL
            ws.cell(r, wc).font = DATA_FONT_B

        # 信心条件字体
        if conf_col:
            v = str(ws.cell(r, conf_col).value or "")
            if "🟢" in v or "较高" in v:
                ws.cell(r, conf_col).font = HIGH_CONF
            elif ("🔴" in v) or ("低" in v and "偏低" not in v):
                ws.cell(r, conf_col).font = LOW_CONF

        # 类型列(胆/双选/全选)条件高亮
        if type_col:
            t = str(ws.cell(r, type_col).value or "").strip()
            if t == "胆":
                ws.cell(r, type_col).fill = BANKER_FILL
                ws.cell(r, type_col).font = BANKER_FONT
            elif t == "双选":
                ws.cell(r, type_col).fill = DOUBLE_FILL
            elif t == "全选":
                ws.cell(r, type_col).fill = TRIPLE_FILL

    # === 底部签名 ===
    sign_row = ws.max_row + 2
    ws.cell(sign_row, 1, value=f"⚡ Claude 神选 · 独立大模型 · 自主推断 · {DATE_STR}")
    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=max_col)
    sc = ws.cell(sign_row, 1)
    sc.font = SIGN_FONT
    sc.fill = SIGN_FILL
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[sign_row].height = 24

    # === 冻结表头(第 3 行起冻结) ===
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last_col_letter}{ws.max_row}"

wb.save(PATH)
print(f"神选 polished: {PATH}")
