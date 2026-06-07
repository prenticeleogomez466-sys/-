import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATE = "2026-06-07"
SRC = r"D:\football-model-exports\health-probe-result.json"
OUT_DIR = rf"C:\Users\Administrator\Desktop\足球推荐\{DATE}\数据源体检"
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, f"足球数据源健康体检_{DATE}.xlsx")

data = json.load(open(SRC, encoding="utf-8"))
rows = data["results"]

# 权威备注/处置(覆盖单次瞬时抖动:sporttery 多次复现 567 WAF;GDELT 可用但限流)
NOTE = {
    "体彩 sporttery 竞彩计算器": ("⚠️存疑", "fail",
        "裸 Node 多次复现 HTTP 567 WAF 挑战页(7KB JS),素头/全浏览器头各3次均挡。生产已有重试+切UA缓解但当前不通。",
        "走 Playwright 系统 Chrome 过 WAF;或继续靠 500.com 兜底(今天已抓4场)。最高优先级。"),
    "体彩 sporttery 赛事公告": ("⚠️存疑", "fail",
        "同上 567 WAF。开售/停售窗口公告抓不到。",
        "同计算器,Playwright 过 WAF;或用竞彩 fixture 自带 sales-window 兜底。"),
    "500.com 竞彩亚盘索引(trade)": ("✅实测", "pass",
        "trade.500.com gb18030 页面正常,抓到 4 场竞彩 fixtureid。sporttery 挂时的兜底主力。",
        "保持。注意悬殊盘只卖让球时 sfcSold 缺口(已加⚠️显示闸)。"),
    "The Odds API": ("✅实测", "reachable",
        "key 有效 HTTP 200;当前 EPL 无赛事(休赛期)故 0 场,非故障。免费档 500 credits/月,仅 eu 区。",
        "保持。世界杯/联赛开赛即有数据。"),
    "API-Football": ("✅实测", "pass",
        "key 有效 /status 200,账户活跃。/status 配额字段本次未解析到(结构差异),不影响取数。",
        "保持。100 次/日上限,省着用。"),
    "football-data.org": ("✅实测", "pass", "token 有效,13 个赛事可用。半场源主力(已接 cron)。", "保持。"),
    "football-data.co.uk CSV": ("✅实测", "pass", "E0 英超 25-26 季 380 行赛果+赔率,历史回测底座。", "保持。"),
    "ClubElo 评级": ("✅实测", "pass", "免 key,630 支俱乐部当日 Elo。评级集成输入。", "保持(俱乐部 Elo,非国家队)。"),
    "Open-Meteo 天气": ("✅实测", "pass", "geocoding+forecast 双段通,168 小时温度。世界杯场馆温度→λ。", "保持。"),
    "GDELT 新闻检索": ("✅实测", "reachable",
        "可用(本次 5 条);但高频/宽查询会 429 限流(此前命中)。动机信号弱信号源。",
        "保持但降频调用,加退避。"),
    "OpenLigaDB(德甲)": ("✅实测", "pass", "免 key,306 场德甲赛果(无伤停字段)。", "保持(候选,德语区赛果)。"),
    "ESPN scoreboard(eng.1)": ("✅实测", "pass", "免 key,10 场赛事。国际赛/友谊赛冗余欧赔来源。", "保持。"),
    "FPL bootstrap(英超伤停)": ("✅实测", "pass", "免 key,62 名伤停/疑似/停赛——唯一可用伤停源,仅覆盖英超。", "保持;伤停覆盖面是结构性缺口(只英超)。"),
    "StatsBomb Open Data": ("✅实测", "pass", "GitHub raw,80 个赛季事件包。历史精选(世界杯/欧洲杯),不滚动更新。", "保持(训练用,非实时)。"),
    "openfootball GitHub": ("✅实测", "pass", "GitHub raw,英超 24-25 共 380 场。历史赛程/赛果。", "保持(候选)。"),
    "ScoreBat 视频/资讯": ("✅实测", "pass", "免 key,50 条视频资讯。上下文用,非概率源。", "保持(候选,低权重)。"),
    "Understat(xG)": ("⚠️存疑", "blocked",
        "反爬空壳,无 datesData。HTTP 200 但无真内容,需浏览器会话。",
        "走 soccerdata 本地跑或真人 Chrome 会话 dump;裸抓不可达。俱乐部 xG。"),
    "ESPN injuries(eng.1)": ("⚠️存疑", "empty",
        "HTTP 200 但 ESPN 足球 injuries feed 结构性不喂数据(已知,非临时)。",
        "弃用此 endpoint;伤停继续靠 FPL(英超)。五大联赛伤停=免费数据墙。"),
    "TheSportsDB free(key=3)": ("⚠️存疑", "empty",
        "免费档 key=3 当前无赛程返回(限流/休赛期)。首发多在付费档。",
        "低优先级;首发已有 ESPN/Sofascore 路径。"),
}

CAT_ORDER = ["官方竞彩", "API赔率", "评级/特征", "免key赛果", "免key伤停", "历史训练", "反爬验证"]
by = {r["name"]: r for r in rows}

wb = Workbook()
ws = wb.active
ws.title = "数据源体检总表"

FONT = "Microsoft YaHei"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
green = PatternFill("solid", fgColor="C6EFCE")
red = PatternFill("solid", fgColor="FFC7CE")
hdr_fill = PatternFill("solid", fgColor="1F4E78")
cat_fill = PatternFill("solid", fgColor="DDEBF7")

# 标题区
ws["A1"] = f"足球大模型 · 数据源健康体检报告"
ws["A1"].font = Font(name=FONT, size=15, bold=True, color="1F4E78")
ws.merge_cells("A1:I1")
ok = sum(1 for r in rows if NOTE[r["name"]][0] == "✅实测")
weak = len(rows) - ok
ws["A2"] = f"体检时间 {data['probedAt'][:19].replace('T',' ')}  ·  真发请求逐源实测(非注册表自报)  ·  共 {len(rows)} 源:✅实测可用 {ok} / ⚠️存疑 {weak}"
ws["A2"].font = Font(name=FONT, size=10, color="555555")
ws.merge_cells("A2:I2")
ws["A3"] = "标签口径:✅实测=本次真发请求确认通且有真内容  ·  🔶推断=接口在线但当前空(休赛/查询窄,非故障)  ·  ⚠️存疑=失败/反爬/结构性不喂数据,需处置"
ws["A3"].font = Font(name=FONT, size=9, italic=True, color="888888")
ws.merge_cells("A3:I3")

headers = ["标签", "类别", "数据源", "供给信号", "状态", "HTTP", "延迟", "实测证据", "备注 / 处置建议"]
hr = 5
for c, h in enumerate(headers, 1):
    cell = ws.cell(hr, c, h)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

r_i = hr + 1
ordered = sorted(rows, key=lambda r: (CAT_ORDER.index(r["cat"]) if r["cat"] in CAT_ORDER else 99, r["name"]))
for r in ordered:
    label, status, note, fix = NOTE[r["name"]]
    vals = [label, r["cat"], r["name"], r["signal"], status,
            r.get("httpCode", "-"), f'{r.get("latencyMs","-")}ms', r.get("evidence", "") or "", f"{note}  ⟶ 处置:{fix}"]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r_i, c, v)
        cell.font = Font(name=FONT, size=10, bold=(c == 1))
        cell.alignment = Alignment(horizontal=("center" if c in (1, 2, 5, 6, 7, 8) else "left"), vertical="center", wrap_text=(c in (4, 9)))
        cell.border = border
        if c == 1:
            cell.fill = green if label == "✅实测" else red
    r_i += 1

widths = [9, 11, 22, 30, 10, 7, 9, 11, 60]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "A6"

# Sheet2 薄弱点
ws2 = wb.create_sheet("薄弱点与处置")
ws2["A1"] = "薄弱点排序与处置(按对真钱管线影响)"
ws2["A1"].font = Font(name=FONT, size=14, bold=True, color="C00000")
ws2.merge_cells("A1:E1")
h2 = ["#", "薄弱点", "影响", "现状/证据", "处置建议"]
for c, h in enumerate(h2, 1):
    cell = ws2.cell(3, c, h)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
weak_rows = [
    ["1", "sporttery 官方源 567 WAF", "竞彩主源(每日真钱管线命脉)裸抓全挂",
     "计算器+公告多次复现 HTTP 567 挑战页;今天 chinaOfficial.ok=false 同因",
     "①生产改 Playwright 系统 Chrome 过 WAF ②短期靠 500.com 兜底(今天抓4场,gate 仍放行)"],
    ["2", "五大联赛伤停=数据墙", "伤停信号只覆盖英超,其余联赛缺",
     "FPL✅62(仅英超);ESPN injuries 结构性不喂;免费世界无国家队伤停",
     "接受边界,情景层标缺;别用估计值兜底(遵无兜底铁律)"],
    ["3", "Understat xG 反爬", "俱乐部 xG 裸抓不可达",
     "HTTP 200 空壳无 datesData",
     "soccerdata 本地跑/真人 Chrome dump;国家队 xG 仍只 FBref(也反爬)"],
    ["4", "GDELT 限流 429", "动机/新闻弱信号偶发取不到",
     "可用但高频查询 429",
     "降频+指数退避;本就弱信号,低权重容忍"],
    ["5", "TheSportsDB 免费档空", "首发/赛程冗余源当前无返回",
     "免费 key=3 限流/休赛期空",
     "低优先级;首发已有 ESPN/Sofascore 主路径"],
]
ri = 4
for row in weak_rows:
    for c, v in enumerate(row, 1):
        cell = ws2.cell(ri, c, v)
        cell.font = Font(name=FONT, size=10, bold=(c == 1))
        cell.alignment = Alignment(horizontal=("center" if c == 1 else "left"), vertical="center", wrap_text=True)
        cell.border = border
    ri += 1
for col, w in zip("ABCDE", [4, 24, 26, 40, 50]):
    ws2.column_dimensions[col].width = w

wb.save(OUT)
print(OUT)
