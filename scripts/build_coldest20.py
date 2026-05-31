# -*- coding: utf-8 -*-
"""每彩种出 20 注【热度最低】号码 (实跑全枚举/采样打分, 可追溯)。
排三/排五全枚举; 大乐透随机采样大池择最低。次级排序用推进指数(直选注数百分位)。
输出桌面 xlsx。注意: 今天周日大乐透不开奖, DLT 为下一开奖期(周一)。"""
import sys, itertools, datetime
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import secrets
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import src.popularity as pop
import crowding_index as ci

TODAY = datetime.date.today().isoformat()
N = 20
model = ci.build_p3_model()

def lucky(s):
    t = []
    if "7" in s: t.append("含7")
    if "8" in s: t.append("含8")
    return "·".join(t) if t else "—"

# ---------- 排三 全枚举 ----------
p3 = []
for d in itertools.product(range(10), repeat=3):
    s = pop.score_digits(d).popularity
    _, idx, _ = ci.estimate_p3(d, model)
    p3.append((s, idx, "".join(map(str, d))))
p3.sort(key=lambda x: (x[0], x[1]))
p3 = p3[:N]

# ---------- 排五 全枚举 (10万) · 热度最低 + 前3位各不相同(分散版) ----------
p5_all = []
for d in itertools.product(range(10), repeat=5):
    s = pop.score_digits(d).popularity
    _, idx, _ = ci.estimate_p3(d[:3], model)   # 推进指数基于前3位真实注数
    p5_all.append((s, idx, "".join(map(str, d))))
p5_all.sort(key=lambda x: (x[0], x[1]))        # 热度↓ 再 推进↓
p5, seen3 = [], set()
for s, idx, c in p5_all:
    if c[:3] in seen3:                          # 前3位去重 -> 分散
        continue
    seen3.add(c[:3]); p5.append((s, idx, c))
    if len(p5) == N:
        break

# ---------- 大乐透 随机采样大池 (今天不开奖 -> 下一期) ----------
rng = secrets.SystemRandom()
pool = {}
for _ in range(400000):
    front = tuple(sorted(rng.sample(range(1, 36), 5)))
    back = tuple(sorted(rng.sample(range(1, 13), 2)))
    key = (front, back)
    if key in pool:
        continue
    s = pop.score_dlt(front, back).popularity
    pool[key] = s
dlt = sorted(((s, f, b) for (f, b), s in pool.items()), key=lambda x: x[0])[:N]

# ---------- 写 xlsx ----------
hdr_fill = PatternFill("solid", fgColor="2F6FED"); hdr_font = Font(bold=True, color="FFFFFF")
title_font = Font(bold=True, size=13, color="1F3864"); note_font = Font(italic=True, size=9, color="808080")
cold_fill = PatternFill("solid", fgColor="DCE9FB")
thin = Side("thin", color="BFBFBF"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
wb = openpyxl.Workbook()

def style_hdr(ws, row, n):
    for j in range(1, n + 1):
        c = ws.cell(row=row, column=j); c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border

def sheet(ws, title, sub, headers, rows):
    ws.cell(1, 1, title).font = title_font
    ws.cell(2, 1, sub).font = note_font
    for j, h in enumerate(headers, 1):
        ws.cell(4, j, h)
    style_hdr(ws, 4, len(headers))
    for i, vals in enumerate(rows, 1):
        for j, v in enumerate([i] + vals, 1):
            c = ws.cell(4 + i, j, v); c.alignment = center; c.border = border
        for j in range(1, len(headers) + 1):
            ws.cell(4 + i, j).fill = cold_fill
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + j)].width = 14

ws1 = wb.active; ws1.title = "排列三"
sheet(ws1, "排列三 第26141期 · 热度最低 20 注 · 40 元",
      f"生成日 {TODAY} · 全1000枚举取最低 · 热度↓推进↓双低 · 非预测",
      ["#", "号码", "热度指数", "推进指数(注数百分位)", "命运数"],
      [[c, s, idx, lucky(c)] for s, idx, c in p3])

ws2 = wb.create_sheet("排列五")
sheet(ws2, "排列五 第26141期 · 热度最低 20 注 · 40 元",
      f"生成日 {TODAY} · 全10万枚举取最低 · 推进指数基于前3位真实注数 · 非预测",
      ["#", "号码", "热度指数", "推进指数(前3位)", "命运数"],
      [[c, s, idx, lucky(c)] for s, idx, c in p5])

ws3 = wb.create_sheet("大乐透(周一26060)")
sheet(ws3, "大乐透 第26060期(周一开奖) · 热度最低 20 注 · 40 元",
      f"生成日 {TODAY} · 40万随机池取最低热度 · 今天周日不开奖 · 单式5+2 · 非预测",
      ["#", "前区", "后区", "热度指数", "命运数"],
      [[" ".join(f"{x:02}" for x in f), " ".join(f"{x:02}" for x in b), s,
        lucky("".join(f"{x:02}" for x in f) + "".join(f"{x:02}" for x in b))]
       for s, f, b in dlt])

out = Path.home() / "Desktop" / f"彩票热度最低20注_{TODAY}.xlsx"
wb.save(out)
print("SAVED:", out)
print("P3 heat range:", p3[0][0], "-", p3[-1][0])
print("P5 heat range:", p5[0][0], "-", p5[-1][0])
print("DLT heat range:", dlt[0][0], "-", dlt[-1][0])
