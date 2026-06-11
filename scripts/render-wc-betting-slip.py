# -*- coding: utf-8 -*-
"""实盘下注单 xlsx 渲染器(神选深紫标准)。
用法: python scripts/render-wc-betting-slip.py <slip.json> <out.xlsx>
Sheet1 实盘下注单(A主力/B小串/C小注) | Sheet2 14场·任选9逐腿裁决 | Sheet3 政策与诚实边界
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PURPLE = "FF4A148C"
GOLD = "FFFFD54F"
RED_TINT = "FFFFEBEE"
GREEN_TINT = "FFE8F5E9"

with open(sys.argv[1], encoding="utf-8") as f:
    S = json.load(f)
OUT = sys.argv[2]

wb = Workbook()
head_font = Font(bold=True, color="FFFFFFFF", size=10.5)
head_fill = PatternFill("solid", fgColor=PURPLE)
banner_font = Font(bold=True, color=PURPLE.replace("FF", "", 1), size=15)
wrap = Alignment(wrap_text=True, vertical="center")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center


def banner(ws, text, ncols, row=1):
    ws.cell(row, 1, text).font = banner_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 26


# ── Sheet1 实盘下注单 ──
ws = wb.active
ws.title = "实盘下注单"
H1 = ["区", "对阵", "开赛", "玩法", "选项", "竞彩赔率", "模型概率", "自评EV", "注金U", f"金额(1U={S['policy']['unit']}元)", "标注(红场保留·临场你定)"]
banner(ws, f"⚡神选·世界杯实盘下注单 · {S['date']} · 决策源=世界杯模型(0611铁律)", len(H1))
ws.cell(2, 1, f"建议总额: 全单{S['totals']['totalU']}U={S['totals']['suggestedAmount']}元 | 核心仓(剔自评EV<-8%){S['totals']['greenU']}U={S['totals']['suggestedCoreAmount']}元 | {S['totals']['note']}")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(H1))
ws.cell(2, 1).font = Font(bold=True, size=10)
ws.row_dimensions[2].height = 30
ws.cell(2, 1).alignment = wrap
ws.append(H1)
style_header(ws, 3, len(H1))
for r in S["rows"]:
    flags = " | ".join(r["flags"]) if r["flags"] else ""
    ws.append([r["section"], r["match"], r["kickoff"], r["play"], r["pick"], r["odds"],
               f"{r['modelProb']*100:.1f}%", f"{r['ev']*100:+.1f}%", r["stakeU"],
               round(r["stakeU"] * S["policy"]["unit"]), flags])
    row = ws.max_row
    tint = RED_TINT if r["ev"] <= -0.08 else (GREEN_TINT if r["ev"] > 0 else None)
    if tint:
        for c in range(1, len(H1) + 1):
            ws.cell(row, c).fill = PatternFill("solid", fgColor=tint)
for p in S["parlays"]:
    ws.append(["B小串", " + ".join(p["legs"]), "", p["name"], "见对阵列", p["combinedOdds"],
               f"{p['combinedProb']*100:.1f}%", f"{p['ev']*100:+.1f}%", p["stakeU"],
               round(p["stakeU"] * S["policy"]["unit"]), "串关=命中率连乘,只串高信心同向腿"])
widths = [13, 30, 12, 12, 12, 9, 9, 9, 7, 9, 46]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for row in ws.iter_rows(min_row=4):
    ws.row_dimensions[row[0].row].height = max(18, 14 * (str(row[10].value or "").count("|") + 1))
    for c in row:
        c.alignment = wrap
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:{get_column_letter(len(H1))}{ws.max_row}"

# ── Sheet2 14场·任选9 ──
F = S.get("fourteen")
ws2 = wb.create_sheet("14场·任选9裁决")
H2 = ["腿", "对阵", "开赛", "主/平/客%", "单选", "复选建议", "胆评级", "防平", "爆冷", "若爆冷后果", "全数据归因"]
period = F.get("period", "") if F else ""
banner(ws2, f"🎯 14场胜负彩 {period} · 逐腿裁决(胆/防平/爆冷) · 世界杯模型全因素", len(H2))
ws2.append(H2)
style_header(ws2, 2, len(H2))
if F and not F.get("error"):
    for l in F["legs"]:
        if l.get("error"):
            ws2.append([l["leg"], l["match"], "", "", "", "", "", "", "", "", l["error"]])
            continue
        pr = l["probs"]
        ws2.append([l["leg"], l["match"], l["kickoff"],
                    f"{pr['home']*100:.0f}/{pr['draw']*100:.0f}/{pr['away']*100:.0f}",
                    l["pick"], l["combo"], l["banker"], l["drawGuard"] or "—",
                    f"{l['upset']}({l['upsetProb']*100:.0f}%·{l.get('upsetDir','')})",
                    l.get("upsetConsequence", "—"), l["reasons"]])
        row = ws2.max_row
        if l["banker"] == "🎯可胆":
            for c in range(1, len(H2) + 1):
                ws2.cell(row, c).fill = PatternFill("solid", fgColor=GREEN_TINT)
        elif l["drawGuard"].startswith("🛡") or l["upset"] == "🔥高":
            for c in range(1, len(H2) + 1):
                ws2.cell(row, c).fill = PatternFill("solid", fgColor=RED_TINT)
    ws2.append([])
    r9 = F["renxuan9"]
    summary = [
        f"🎯 可胆腿: {','.join(map(str, F['bankers'])) or '无'}",
        f"🛡 必防平腿: {','.join(map(str, F['drawGuards'])) or '无'}",
        f"🔥 高爆冷腿: {','.join(map(str, F['upsetWatch'])) or '无'}",
        f"任选9: {' '.join(r9['picks'])} | 9腿全中概率={r9['combinedProb']*100:.2f}%({r9['note']})",
    ]
    if r9.get("riskLegs"):
        summary.append(f"⚠️ 任选9命门腿(冷向概率最高,炸一腿票即死): {' ║ '.join(r9['riskLegs'])}")
    U = F.get("upsetScenario")
    if U:
        st, ct = U["singleTicket"], U["comboTicket"]
        summary.append("")
        summary.append("💣 爆冷情景推演(" + U["assumption"] + ")")
        summary.append(f"  单选票: 全中率={st['pAllHit']*100:.3f}% · ≥13中={st['pAtLeast13']*100:.2f}% · 期望命中={st['expHits']}腿({st['note']})")
        summary.append(f"  复选票(按建议买): 全中率={ct['pAllHit']*100:.2f}% · {ct['tickets']}注={ct['costYuan']}元({ct['note']})")
        for t in U["topUpsetLegs"]:
            summary.append(f"  🔥第{t['leg']}腿 {t['match']} | 冷向={t['dir']}({t['prob']*100:.0f}%) | 依据: {t['reasons']} | 后果: {t['consequence']}")
        for p in U["topColdPairs"]:
            summary.append(f"  💥双冷组合: 第{p['legs'][0]}+{p['legs'][1]}腿({p['matches']}) 联合概率{p['prob']*100:.1f}%")
    for s in summary:
        ws2.append([s])
        ws2.merge_cells(start_row=ws2.max_row, start_column=1, end_row=ws2.max_row, end_column=len(H2))
        ws2.cell(ws2.max_row, 1).font = Font(bold=True, size=10.5)
w2 = [5, 26, 12, 13, 10, 16, 10, 17, 14, 30, 46]
for i, w in enumerate(w2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for row in ws2.iter_rows(min_row=3):
    ws2.row_dimensions[row[0].row].height = 20
    for c in row:
        c.alignment = wrap
ws2.freeze_panes = "A3"

# ── Sheet3 政策与诚实边界 ──
ws3 = wb.create_sheet("政策与诚实边界")
banner(ws3, "📜 四项裁决(2026-06-11你定的)+诚实边界+本单跳过项", 2)
rows3 = [
    ("预算", S["policy"]["budget"]),
    ("玩法", S["policy"]["plays"]),
    ("红场", S["policy"]["redHandling"]),
    ("注金", S["policy"]["staking"]),
    ("决策源", f"{S['source']['model']}(预测{S['source']['predictionsAt']}) · 竞彩赔率快照{S['source']['jingcaiOddsAt']}"),
    ("诚实边界", S["honesty"]),
]
for k, v in rows3:
    ws3.append([k, v])
    ws3.cell(ws3.max_row, 1).font = Font(bold=True)
ws3.append([])
ws3.append(["跳过项(全部如实记录,绝不硬凑)", ""])
ws3.cell(ws3.max_row, 1).font = Font(bold=True, color=PURPLE.replace("FF", "", 1))
for s in S["skipped"]:
    ws3.append(["", s])
ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 110
for row in ws3.iter_rows(min_row=2):
    for c in row:
        c.alignment = wrap

wb.save(OUT)
print("saved", OUT)
