"""
2026-05-28 竞彩 5 场推荐 XLSX(一致性约束版).

核心原则:**比分是锚点,胜负平/让球方向/半全场都从比分推导,保证逻辑一致**.
独立的让球赔率分析仍展示作为"市场分歧 view".
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from itertools import combinations

OUTPUT = r"C:\Users\Administrator\Desktop\2026-05-28 竞彩足球推荐.xlsx"

# ───── 5 场原始数据 + 人工综合分析(model_score 是核心,其他从此推导)─────
fixtures = [
    {
        "seq": "001", "comp": "国际赛", "kickoff": "05-29 02:45",
        "home": "爱尔兰", "away": "卡塔尔",
        "odds0": (1.40, 3.88, 6.35),
        "h_line": -1, "oddsH": (2.50, 3.07, 2.48),
        "hf": {"胜胜":2.02,"胜平":18.00,"胜负":55.00,"平胜":3.90,"平平":5.90,"平负":12.50,"负胜":27.00,"负平":18.00,"负负":11.50},
        "all_scores": {"1-0":5.50,"2-0":6.00,"2-1":6.80,"3-0":10.00,"3-1":11.00,"0-0":11.50,"1-1":7.00,"2-2":17.00,"0-1":14.50,"0-2":35.00,"1-2":18.00},
        "comp_type": "友谊赛",
        "lineup_note": "国际友谊赛,两队均可能轮换;爱尔兰主场实力略优",
        # 锚点:模型推荐的比分(从赔率+让球+阵容综合)
        "model_score": "1-0",
        "score_odds": 5.50,
        "score_reason": "让 -1 后客胜 2.48 最低 → 庄家觉得最多赢 1 球;比分 1-0 (5.50) 是赢 1 球里最低赔率",
        "model_score_alt": "2-1",
        "score_alt_odds": 6.80,
        "lineup_recommend": "可选",
        "upset_level": "高",
        "upset_reason": "友谊赛 + 卡塔尔实力接近(60 vs 48 FIFA 排名);1.40 没安全垫",
    },
    {
        "seq": "002", "comp": "葡超", "kickoff": "05-29 03:00",
        "home": "卡萨皮亚", "away": "托林斯",
        "odds0": (2.11, 2.73, 3.46),
        "h_line": -1, "oddsH": (5.40, 3.40, 1.54),
        "hf": {"胜胜":4.00,"胜平":15.00,"胜负":33.00,"平胜":4.15,"平平":3.85,"平负":6.75,"负胜":30.00,"负平":15.00,"负负":6.30},
        "all_scores": {"1-0":5.40,"2-0":8.50,"2-1":7.50,"0-0":7.00,"1-1":5.25,"2-2":17.00,"0-1":7.25,"0-2":14.00,"1-2":10.50},
        "comp_type": "联赛",
        "lineup_note": "葡超低位附加赛;让 -1 主队跳到 5.40",
        "model_score": "1-1",
        "score_odds": 5.25,
        "score_reason": "半全场平平 3.85 整场最低 + 比分 1-1 (5.25) 整场最低 → 庄家真实预期平局",
        "model_score_alt": "0-0",
        "score_alt_odds": 7.00,
        "lineup_recommend": "强推",
        "upset_level": "中",
        "upset_reason": "主胜赔 2.11 是诱饵,让 -1 后崩到 5.40;走平局是顺势",
    },
    {
        "seq": "003", "comp": "解放者杯", "kickoff": "05-29 06:00",
        "home": "波特诺", "away": "水晶体育",
        "odds0": (1.51, 3.40, 5.80),
        "h_line": -1, "oddsH": (3.05, 2.85, 2.22),
        "hf": {"胜胜":2.40,"胜平":17.00,"胜负":55.00,"平胜":3.50,"平平":5.00,"平负":12.50,"负胜":23.00,"负平":17.00,"负负":11.00},
        "all_scores": {"1-0":5.50,"2-0":6.50,"2-1":6.75,"3-0":12.00,"0-0":10.00,"1-1":6.70,"2-2":17.00,"0-1":11.00,"0-2":26.00,"1-2":14.50},
        "comp_type": "杯赛-小组",
        "lineup_note": "解放者杯小组赛;双方无重大伤停",
        "model_score": "1-0",
        "score_odds": 5.50,
        "score_reason": "让 -1 客胜 2.22 最低 → 庄家觉得主队最多 1 球优势;1-0 是赢 1 球里最低赔率",
        "model_score_alt": "2-1",
        "score_alt_odds": 6.75,
        "lineup_recommend": "可选",
        "upset_level": "中",
        "upset_reason": "杯赛客队反扑历史多发;让球后没安全垫",
    },
    {
        "seq": "004", "comp": "解放者杯", "kickoff": "05-29 06:00",
        "home": "帕梅拉斯", "away": "巴兰基亚",
        "odds0": (1.14, 5.95, 12.00),
        "h_line": -2, "oddsH": (2.85, 3.45, 2.05),
        "hf": {"胜胜":1.50,"胜平":30.00,"胜负":90.00,"平胜":3.70,"平平":9.20,"平负":23.00,"负胜":24.00,"负平":30.00,"负负":22.00},
        "all_scores": {"1-0":6.50,"2-0":5.00,"2-1":8.50,"3-0":5.50,"3-1":10.00,"4-0":10.00,"0-0":17.00,"1-1":11.00,"0-1":30.00,"0-2":85.00,"1-2":38.00},
        "comp_type": "杯赛-小组",
        "lineup_note": "巴兰基亚队长查拉缺席 + 中卫蒙松出战成疑;帕梅主力齐",
        "model_score": "2-0",
        "score_odds": 5.00,
        "score_reason": "让 -2 主胜仍 2.85 → 庄家认为赢 2 球+正常;比分 2-0 (5.00) 整场最低",
        "model_score_alt": "3-0",
        "score_alt_odds": 5.50,
        "lineup_recommend": "强推",
        "upset_level": "低",
        "upset_reason": "唯一风险:帕梅小组出线后出工不出力 → 平局风险",
    },
    {
        "seq": "005", "comp": "解放者杯", "kickoff": "05-29 08:30",
        "home": "博卡", "away": "天主大学",
        "odds0": (1.36, 4.00, 6.95),
        "h_line": -1, "oddsH": (2.53, 2.80, 2.65),
        "hf": {"胜胜":1.95,"胜平":22.00,"胜负":65.00,"平胜":3.55,"平平":6.00,"平负":14.00,"负胜":25.00,"负平":22.00,"负负":13.50},
        "all_scores": {"1-0":5.00,"2-0":5.00,"2-1":7.75,"3-0":8.00,"0-0":10.00,"1-1":8.00,"2-2":25.00,"0-1":15.00,"0-2":40.00,"1-2":22.00},
        "comp_type": "杯赛-小组",
        "lineup_note": "博卡 3 主力伤(卡瓦尼+基万尼斯+阿斯卡西巴尔);锋线受损",
        "model_score": "1-0",
        "score_odds": 5.00,
        "score_reason": "让 -1 后三档接近三平 → 庄家不信博卡赢超 1 球;阵容残缺锋线无力,1-0 比 2-0 更稳",
        "model_score_alt": "2-0",
        "score_alt_odds": 5.00,
        "lineup_recommend": "谨慎",
        "upset_level": "中高",
        "upset_reason": "3 主力伤 + 让 -1 后接近三平 → 平局或客胜爆冷概率显著高于赔率隐含",
    }
]

# ───── 一致性推导函数 ─────

def parse_score(score_str):
    h, a = score_str.split("-")
    return int(h), int(a)

def wld_from_score(home_goals, away_goals):
    """从比分推 胜负平方向"""
    if home_goals > away_goals: return "主胜"
    if home_goals < away_goals: return "客胜"
    return "平局"

def handicap_outcome_from_score(home_goals, away_goals, line):
    """让 line(负值=主队让N球)后的胜平负 outcome"""
    adjusted_diff = (home_goals + line) - away_goals
    if adjusted_diff > 0: return "主胜"
    if adjusted_diff < 0: return "客胜"
    return "平局"

def derive_consistent_halffull(f, score_str, target_full_outcome):
    """挑跟比分一致的最低赔率半全场.

    比分 X-Y → 全场 outcome (主胜/平/客胜)
    半全场必须满足:
      1. 全场字符 = target_full_outcome 对应字符(胜/平/负)
      2. 上半场结果 + 下半场结果能拼出 X-Y(理论上半场 0-0/部分加和都可能)

    简化:筛 endswith target_char,挑最低赔率
    """
    h, a = parse_score(score_str)
    target_char = "胜" if target_full_outcome == "主胜" else "平" if target_full_outcome == "平局" else "负"
    candidates = {k: v for k, v in f["hf"].items() if k.endswith(target_char)}
    # 进一步:对每个候选,检查上半场结果是否能拼到全场(简单 plausibility 检查)
    plausible = {}
    for hf_label, odds in candidates.items():
        first_char = hf_label[0]
        if score_plausible_first_half(first_char, h, a):
            plausible[hf_label] = odds
    pool = plausible if plausible else candidates
    sorted_pool = sorted(pool.items(), key=lambda x: x[1])
    return sorted_pool[0] if sorted_pool else (None, None)

def score_plausible_first_half(first_char, full_h, full_a):
    """上半场字符是否可能加到全场 full_h-full_a.
    上半场字符 胜=主进>客进, 平=主进==客进, 负=主进<客进
    存在 h1 <= full_h, a1 <= full_a 使条件满足?
    """
    if first_char == "胜":
        # 需要 h1>a1 with h1<=full_h, a1<=full_a
        # 只要 full_h >= 1 就有可能(h1=1, a1=0)
        return full_h >= 1
    if first_char == "平":
        # 需要 h1==a1
        # 0-0 都可以
        return True
    # 负:需要 h1<a1
    return full_a >= 1

def pick_consistent_handicap_outcome(score_str, h_line):
    """让球玩法的方向(从比分推)"""
    h, a = parse_score(score_str)
    return handicap_outcome_from_score(h, a, h_line)

def pick_independent_handicap(odds_h):
    """让球独立分析:挑赔率最低的方向(可能跟比分推导不一致 → 标"市场分歧")"""
    candidates = [("主胜", odds_h[0]), ("平局", odds_h[1]), ("客胜", odds_h[2])]
    return min(candidates, key=lambda x: x[1])

# ───── 给每场跑一致性推导 ─────
for f in fixtures:
    # 1. 主比分 → 全场胜负平
    f["wld"] = wld_from_score(*parse_score(f["model_score"]))
    # 2. 比分 → 让球玩法方向(consistent view)
    f["handicap_consistent"] = pick_consistent_handicap_outcome(f["model_score"], f["h_line"])
    f["handicap_consistent_odds"] = f["oddsH"][["主胜", "平局", "客胜"].index(f["handicap_consistent"])]
    # 3. 比分 → 一致半全场
    hf_label, hf_odds = derive_consistent_halffull(f, f["model_score"], f["wld"])
    f["hf_consistent"] = hf_label
    f["hf_consistent_odds"] = hf_odds
    # 4. 替代半全场:剩余候选里第二低
    target_char = "胜" if f["wld"] == "主胜" else "平" if f["wld"] == "平局" else "负"
    rest_hf = sorted([(k, v) for k, v in f["hf"].items() if k.endswith(target_char) and k != hf_label], key=lambda x: x[1])
    f["hf_alt"] = rest_hf[0] if rest_hf else (None, None)
    # 5. 让球独立分析(可能跟 consistent 不同)
    f["handicap_independent_label"], f["handicap_independent_odds"] = pick_independent_handicap(f["oddsH"])
    f["handicap_market_split"] = (f["handicap_independent_label"] != f["handicap_consistent"])

# ───── 样式 ─────
TITLE_FONT = Font(name="微软雅黑", bold=True, size=15, color="FFFFFF")
TITLE_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="305496")
SECTION_FONT = Font(name="微软雅黑", bold=True, size=13, color="FFFFFF")
SECTION_FILL = PatternFill("solid", start_color="2E75B6")
DEFAULT_FONT = Font(name="微软雅黑", size=11)
BOLD_FONT = Font(name="微软雅黑", bold=True, size=11)
RECOMMEND_FONT = Font(name="微软雅黑", bold=True, color="C00000", size=11)
ALT_FILL = PatternFill("solid", start_color="F2F7FC")
WARN_FILL = PatternFill("solid", start_color="FFC7CE")
THIN = Side(border_style="thin", color="9CC3E5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

DIR_COLORS = {"主胜": "A9D08E", "平局": "FFD966", "客胜": "F4B084"}
REC_COLORS = {
    "强推": ("00B050", "FFFFFF"), "推荐": ("92D050", "000000"), "可选": ("FFC000", "000000"),
    "谨慎": ("ED7D31", "FFFFFF"), "避坑": ("C00000", "FFFFFF")
}
UPSET_COLORS = {"低": "C6E0B4", "中": "FFE699", "中高": "F8CBAD", "高": "F4B084"}

def hcell(cell):
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER

def scell(cell):
    cell.fill = SECTION_FILL; cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = BORDER

def dcell(cell, alt=False, bold=False, rec=False, wrap=False):
    cell.font = RECOMMEND_FONT if rec else (BOLD_FONT if bold else DEFAULT_FONT)
    cell.alignment = LEFT_WRAP if wrap else CENTER
    cell.border = BORDER
    if alt and cell.fill.start_color.rgb is None: cell.fill = ALT_FILL

# ───── workbook ─────
wb = Workbook()
s = wb.active
s.title = "2026-05-28 竞彩推荐"

# 标题
s.cell(row=1, column=1, value="🏆 2026-05-28 竞彩 5 场推荐(逻辑一致版:比分锚点,其余从比分推导)")
s.cell(row=1, column=1).font = TITLE_FONT
s.cell(row=1, column=1).fill = TITLE_FILL
s.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
s.merge_cells("A1:M1")
s.row_dimensions[1].height = 36

s.cell(row=2, column=1, value="✓ 比分 = 综合赔率+让球盘+半全场赔率+阵容判断  →  胜负平 / 让球方向 / 半全场 全部从比分推导,绝对一致")
s.cell(row=2, column=1).font = Font(name="微软雅黑", italic=True, size=11, color="C00000")
s.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
s.merge_cells("A2:M2")
s.row_dimensions[2].height = 22

# 区段 1: 5 场推荐(一致性)
s.cell(row=4, column=1, value="一、5 场推荐(比分锚点 → 全部一致)")
scell(s.cell(row=4, column=1))
s.merge_cells("A4:M4")
s.row_dimensions[4].height = 26

headers = [
    "编号", "联赛", "性质", "对阵",
    "**比分首选**", "**比分备选**",
    "胜负平方向", "让 0 赔率",
    "让球方向\n(一致)", "让球赔率",
    "**半全场首选**", "**半全场备选**",
    "建议"
]
for j, h in enumerate(headers, 1):
    hcell(s.cell(row=5, column=j, value=h))
s.row_dimensions[5].height = 40

for i, f in enumerate(fixtures, 6):
    alt = (i % 2 == 0)
    wld = f["wld"]
    s.cell(row=i, column=1, value=f"周四{f['seq']}")
    s.cell(row=i, column=2, value=f["comp"])
    s.cell(row=i, column=3, value=f["comp_type"])
    s.cell(row=i, column=4, value=f"{f['home']}\nVS\n{f['away']}")
    s.cell(row=i, column=5, value=f"{f['model_score']}\n(赔 {f['score_odds']})")
    s.cell(row=i, column=6, value=f"{f['model_score_alt']}\n(赔 {f['score_alt_odds']})")
    s.cell(row=i, column=7, value=wld)
    s.cell(row=i, column=8, value=f["odds0"][["主胜", "平局", "客胜"].index(wld)])
    handicap_label = f"让{f['h_line']}\n{f['handicap_consistent']}"
    s.cell(row=i, column=9, value=handicap_label)
    s.cell(row=i, column=10, value=f["handicap_consistent_odds"])
    s.cell(row=i, column=11, value=f"{f['hf_consistent']}\n(赔 {f['hf_consistent_odds']})" if f["hf_consistent"] else "—")
    s.cell(row=i, column=12, value=f"{f['hf_alt'][0]}\n(赔 {f['hf_alt'][1]})" if f["hf_alt"][0] else "—")
    s.cell(row=i, column=13, value=f["lineup_recommend"])

    for j in range(1, 14):
        c = s.cell(row=i, column=j)
        is_rec = j in [5, 6, 11, 12]
        dcell(c, alt=alt, bold=(j == 4), rec=is_rec)
    # 颜色
    s.cell(row=i, column=7).fill = PatternFill("solid", start_color=DIR_COLORS.get(wld, "FFFFFF"))
    s.cell(row=i, column=9).fill = PatternFill("solid", start_color=DIR_COLORS.get(f["handicap_consistent"], "FFFFFF"))
    rc, rt = REC_COLORS.get(f["lineup_recommend"], ("808080", "FFFFFF"))
    s.cell(row=i, column=13).fill = PatternFill("solid", start_color=rc)
    s.cell(row=i, column=13).font = Font(name="微软雅黑", bold=True, color=rt, size=11)
    s.cell(row=i, column=8).number_format = "0.00"
    s.cell(row=i, column=10).number_format = "0.00"
    s.row_dimensions[i].height = 62

# 区段 2: 一致性验证 + 让球市场分歧 view
row = 12
s.cell(row=row, column=1, value="二、一致性验证 + 让球市场分歧 view")
scell(s.cell(row=row, column=1))
s.merge_cells(f"A{row}:M{row}")
s.row_dimensions[row].height = 24
row += 1

sub_headers = ["编号", "对阵", "比分", "上半场字符\n(从半全场首选)", "下半场\n推导", "全场胜负平\n(从比分)", "让球玩法\n一致 view", "让球玩法\n独立赔率最低", "市场分歧?", "解读"]
for j, h in enumerate(sub_headers, 1):
    hcell(s.cell(row=row, column=j, value=h))
s.merge_cells(start_row=row, start_column=10, end_row=row, end_column=13)
hcell(s.cell(row=row, column=10, value="解读 / 市场结构提示"))
s.row_dimensions[row].height = 38
row += 1

for f in fixtures:
    alt = (row % 2 == 0)
    h, a = parse_score(f["model_score"])
    half_first_char = f["hf_consistent"][0] if f["hf_consistent"] else "?"
    # 推导下半场
    if half_first_char == "胜":
        second_half_label = f"主队下半场再 {h - 1}-{a} 或类似"
    elif half_first_char == "平":
        second_half_label = f"下半场主队进 {h} 客队进 {a}"
    else:
        second_half_label = f"上半场客队领先,下半场主队反超到 {h}-{a}"

    s.cell(row=row, column=1, value=f"周四{f['seq']}")
    s.cell(row=row, column=2, value=f"{f['home']}\nVS {f['away']}")
    s.cell(row=row, column=3, value=f["model_score"])
    s.cell(row=row, column=4, value=half_first_char)
    s.cell(row=row, column=5, value=second_half_label)
    s.cell(row=row, column=6, value=f["wld"])
    s.cell(row=row, column=7, value=f["handicap_consistent"])
    s.cell(row=row, column=8, value=f["handicap_independent_label"])
    s.cell(row=row, column=9, value="⚠ 分歧" if f["handicap_market_split"] else "✓ 一致")

    if f["handicap_market_split"]:
        split_note = (f"市场对让 {f['h_line']} 最看好「{f['handicap_independent_label']}」(赔 {f['handicap_independent_odds']}),"
                      f"但我们的比分 {f['model_score']} 推出让 {f['h_line']} 应该走「{f['handicap_consistent']}」(赔 {f['handicap_consistent_odds']})。"
                      "这是双轨建议:跟比分一起买就走一致 view;单独投让球玩法可参考独立 view")
    else:
        split_note = f"市场让球赔率最低方向跟比分推导一致 → 让球玩法可放心跟随"
    s.cell(row=row, column=10, value=split_note)
    s.merge_cells(start_row=row, start_column=10, end_row=row, end_column=13)

    for j in range(1, 14):
        c = s.cell(row=row, column=j)
        dcell(c, alt=alt, wrap=(j in [2, 5, 10]))
    s.cell(row=row, column=6).fill = PatternFill("solid", start_color=DIR_COLORS.get(f["wld"], "FFFFFF"))
    s.cell(row=row, column=7).fill = PatternFill("solid", start_color=DIR_COLORS.get(f["handicap_consistent"], "FFFFFF"))
    s.cell(row=row, column=8).fill = PatternFill("solid", start_color=DIR_COLORS.get(f["handicap_independent_label"], "FFFFFF"))
    if f["handicap_market_split"]:
        s.cell(row=row, column=9).fill = WARN_FILL
    s.row_dimensions[row].height = 80
    row += 1

# 区段 3: 深度分析
row += 1
s.cell(row=row, column=1, value="三、深度分析(为什么选这个比分)")
scell(s.cell(row=row, column=1))
s.merge_cells(f"A{row}:M{row}")
s.row_dimensions[row].height = 24
row += 1

for j, h in enumerate(["编号", "对阵", "比分理由", "阵容/伤停", "爆冷"], 1):
    hcell(s.cell(row=row, column=j, value=h))
s.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
hcell(s.cell(row=row, column=3, value="比分理由"))
s.merge_cells(start_row=row, start_column=5, end_row=row, end_column=10)
hcell(s.cell(row=row, column=5, value="阵容/伤停"))
s.merge_cells(start_row=row, start_column=11, end_row=row, end_column=13)
hcell(s.cell(row=row, column=11, value="爆冷分析"))
s.row_dimensions[row].height = 28
row += 1

for f in fixtures:
    alt = (row % 2 == 0)
    s.cell(row=row, column=1, value=f"周四{f['seq']}")
    s.cell(row=row, column=2, value=f"{f['home']}\nVS {f['away']}")
    s.cell(row=row, column=3, value=f["score_reason"])
    s.cell(row=row, column=5, value=f["lineup_note"])
    s.cell(row=row, column=11, value=f"{f['upset_level']}: {f['upset_reason']}")
    for j in range(1, 14):
        c = s.cell(row=row, column=j)
        dcell(c, alt=alt, wrap=(j >= 2))
    s.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    s.merge_cells(start_row=row, start_column=5, end_row=row, end_column=10)
    s.merge_cells(start_row=row, start_column=11, end_row=row, end_column=13)
    s.row_dimensions[row].height = 56
    row += 1

# 区段 4: 串关组合(基于一致 view)
row += 1
s.cell(row=row, column=1, value="四、串关组合(基于「比分推导出的胜负平」)")
scell(s.cell(row=row, column=1))
s.merge_cells(f"A{row}:M{row}")
s.row_dimensions[row].height = 24
row += 1

combo_headers = ["类型", "组合内容", "联合赔率", "联合概率", "联合 EV", "10 元回报", "10 元净盈", "建议"]
for j, h in enumerate(combo_headers, 1):
    hcell(s.cell(row=row, column=j, value=h))
s.row_dimensions[row].height = 24
row += 1

def odds_for_wld(f, wld):
    return f["odds0"][["主胜", "平局", "客胜"].index(wld)]

def normalize(odds):
    inv = [1/o for o in odds]
    t = sum(inv)
    return tuple(p/t for p in inv)

def prob_for_wld(f, wld):
    p = normalize(f["odds0"])
    return p[["主胜", "平局", "客胜"].index(wld)]

legs = []
for f in fixtures:
    legs.append({
        "seq": f["seq"], "home": f["home"], "away": f["away"],
        "wld": f["wld"], "odds": odds_for_wld(f, f["wld"]), "prob": prob_for_wld(f, f["wld"]),
        "rec": f["lineup_recommend"]
    })

def fmt(l):
    if l["wld"] == "主胜": return f"#{l['seq']}{l['home']}胜"
    if l["wld"] == "客胜": return f"#{l['seq']}{l['away']}胜"
    return f"#{l['seq']}{l['home']}-{l['away']}平"

def stats(combo):
    co = 1; cp = 1
    for l in combo:
        co *= l["odds"]; cp *= l["prob"]
    return co, cp, cp * co - 1

def advice(combo, ev):
    if any(l["rec"] == "避坑" for l in combo): return "不建议(含避坑)"
    if all(l["rec"] in ["强推"] for l in combo): return "可考虑(全强推)"
    if any(l["rec"] == "谨慎" for l in combo) and len(combo) >= 3: return "谨慎(多腿+含谨慎)"
    if ev < -0.35: return "EV 太负"
    return "可考虑"

def fill_combo(r, label, combo_list, alt=False, bold=False):
    co, cp, ev = stats(combo_list)
    s.cell(row=r, column=1, value=label)
    s.cell(row=r, column=2, value=" + ".join(fmt(l) for l in combo_list))
    s.cell(row=r, column=3, value=round(co, 3))
    s.cell(row=r, column=4, value=round(cp, 4))
    s.cell(row=r, column=5, value=round(ev, 4))
    s.cell(row=r, column=6, value=round(10 * co, 2))
    s.cell(row=r, column=7, value=round(10 * co - 10, 2))
    s.cell(row=r, column=8, value=advice(combo_list, ev))
    for j in range(1, 9):
        c = s.cell(row=r, column=j)
        dcell(c, alt=alt, bold=bold, rec=(j == 8))
    s.cell(row=r, column=3).number_format = "0.00"
    s.cell(row=r, column=4).number_format = "0.0%"
    s.cell(row=r, column=5).number_format = "0.0%;[Red](0.0%)"
    s.cell(row=r, column=6).number_format = "0.00"
    s.cell(row=r, column=7).number_format = "0.00"
    s.row_dimensions[r].height = 24

# 五串一
fill_combo(row, "五串一", legs, bold=True); row += 1
# 四串一 top 3
four = sorted([(stats(c), c) for c in combinations(legs, 4)], key=lambda x: x[0][2], reverse=True)[:3]
for k, (_, combo) in enumerate(four):
    fill_combo(row, f"四串一 #{k+1}", combo, alt=(row % 2 == 0)); row += 1
# 三串一 top 5
three = sorted([(stats(c), c) for c in combinations(legs, 3)], key=lambda x: x[0][2], reverse=True)[:5]
for k, (_, combo) in enumerate(three):
    fill_combo(row, f"三串一 #{k+1}", combo, alt=(row % 2 == 0)); row += 1
# 二串一 top 5
two = sorted([(stats(c), c) for c in combinations(legs, 2)], key=lambda x: x[0][2], reverse=True)[:5]
for k, (_, combo) in enumerate(two):
    fill_combo(row, f"二串一 #{k+1}", combo, alt=(row % 2 == 0)); row += 1

# 备注
row += 1
s.cell(row=row, column=1, value="重要说明").font = Font(name="微软雅黑", bold=True, size=12, color="C00000")
for txt in [
    "✓ 一致性保证:比分 → 胜负平 / 让球方向 / 半全场,全部从同一个比分推导,逻辑链条无矛盾",
    "✓ 让球分歧 view:如果让球独立赔率最低的方向 ≠ 我们比分推出的方向,标「⚠ 分歧」 — 此时让球玩法可独立选,但跟比分一起买必须用一致 view",
    "✓ 示例 #005:博卡比分 1-0,让 -1 后博卡赢 0 球 → 让球玩法是「平局」(2.80);市场让 -1 主胜赔率 2.53 最低 → 分歧;两种买法都对,但不能混用",
    "✓ 示例 #004:帕梅 2-0,让 -2 后 0-0 → 让球玩法「平局」(3.45);市场让 -2 客胜 2.05 最低 → 分歧;独立投让球可走客胜博中等赔率",
    "✓ 半全场半场字符 = 胜/平/负 必须能拼到全场比分(plausibility check 已加)",
    "胜负彩本质长期 EV 为负,小额娱乐为主"
]:
    row += 1
    s.cell(row=row, column=1, value=txt).font = DEFAULT_FONT
    s.cell(row=row, column=1).alignment = LEFT_WRAP
    s.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    s.row_dimensions[row].height = 22

# 列宽
widths = {"A": 9, "B": 11, "C": 11, "D": 14, "E": 14, "F": 13, "G": 11, "H": 9, "I": 11, "J": 10, "K": 14, "L": 14, "M": 11}
for letter, w in widths.items():
    s.column_dimensions[letter].width = w

s.freeze_panes = "A6"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print()
print("=== 一致性验证 ===")
for f in fixtures:
    print(f"#{f['seq']} {f['home']:6} VS {f['away']:6} | 比分 {f['model_score']} → 胜负平 {f['wld']} → 让{f['h_line']} {f['handicap_consistent']}({f['handicap_consistent_odds']}) → 半全场 {f['hf_consistent']}({f['hf_consistent_odds']}) | 市场独立让球 {f['handicap_independent_label']}({f['handicap_independent_odds']}) {'⚠分歧' if f['handicap_market_split'] else '✓一致'}")
